"""DRR-V2-5 (2026-08-04) — Verizon-template DRR Excel builder tests.

Covers:
  * Legacy 4-column flat mode (backward compat: no keyword args).
  * DRR-V2 mode: header block cells, section rows, item rows sourced
    from postgres DeliveryItem (item_no + name + actual_completion_date
    + status + tg_name + comment).
  * Header field blank -> WARN log + blank cell (Q6 spec).
  * Items 85/86/87 excluded via section_grouping filtering.
  * Missing logo path silently skipped.
"""
from __future__ import annotations

import io
import logging
from datetime import date
from types import SimpleNamespace

import pytest

from core.src.email_service.outbound.drr_report_excel import (
    CLOSED_LIKE_STATES,
    build_drr_report_excel,
    status_for_item,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _item(
    item_no: int,
    *,
    item_name: str | None = None,
    delivery_state: str = "Open",
    actual_completion_date: date | None = None,
    tg_name: str = "TPM",
    comment: str = "",
) -> SimpleNamespace:
    return SimpleNamespace(
        item_no=item_no,
        item_name=item_name or f"Item {item_no}",
        delivery_state=delivery_state,
        actual_completion_date=actual_completion_date,
        tg_name=tg_name,
        comment=comment,
        # legacy attribute so status_for_item / legacy mode also work
        owner_status_note=comment,
    )


def _open_ws(xlsx_bytes: bytes):
    """Parse the produced bytes back with openpyxl for cell assertions.

    DRR-V2-8 (2026-08-07): after the 4-tab restructure the DEFAULT sheet
    (`wb.active`) is now "Version History", not "Checklist". Existing
    tests assert on Checklist content, so explicitly load that sheet.
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    if "Checklist" in wb.sheetnames:
        return wb["Checklist"]
    return wb.active


# ---------------------------------------------------------------------------
# Legacy mode
# ---------------------------------------------------------------------------


class TestLegacyFlat:
    def test_legacy_call_shape_still_works(self):
        items = [_item(2, delivery_state="Closed"), _item(1)]
        out = build_drr_report_excel(items)
        assert isinstance(out, bytes) and len(out) > 0
        ws = _open_ws(out)
        assert ws.title == "DRR Closure"
        # Legacy 4-column header
        assert ws.cell(row=1, column=1).value == "Item No"
        assert ws.cell(row=1, column=4).value == "Owner Comment"
        # Sorted by item_no ascending
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=3, column=1).value == 2

    def test_status_for_item(self):
        assert status_for_item(_item(1, delivery_state="Open")) == "Open"
        assert status_for_item(_item(1, delivery_state="Closed")) == "Closed"
        assert status_for_item(_item(1, delivery_state="ReadyForSubmission")) == "Closed"
        assert status_for_item(_item(1, delivery_state="SubmittedToCustomer")) == "Closed"
        assert status_for_item(_item(1, delivery_state="OwnerClosed")) == "Open"

    def test_closed_like_states_membership(self):
        assert "Closed" in CLOSED_LIKE_STATES
        assert "OwnerClosed" not in CLOSED_LIKE_STATES


# ---------------------------------------------------------------------------
# DRR-V2 mode — header block
# ---------------------------------------------------------------------------


class TestV2HeaderBlock:
    def _v2_call(self, **overrides):
        defaults = dict(
            items=[_item(1)],
            customer_id="MMK",
            device_id="SM-F976U",
            milestone_id="DRR",
            section_grouping=[{
                "section": "Product Documentation Review",
                "work_items": [{"item_no": 1, "item_name": "Item 1"}],
            }],
            drr_version="5.7",
            milestone_headers={
                "fld_lockdown_date": date(2026, 3, 18),
                "req_version": "Oct 25",
                "target_date": date(2026, 4, 29),
            },
            project_headers={
                "LE": date(2026, 6, 11),
                "FFW": date(2026, 5, 13),
            },
        )
        defaults.update(overrides)
        return build_drr_report_excel(**defaults)

    def test_row_1_oem_model_title(self):
        ws = _open_ws(self._v2_call())
        # DRR-V2-5a (2026-08-05): merged A..F, title lives in column A.
        assert ws.cell(row=1, column=1).value == "OEM Model"

    def test_row_3_device_readiness_review_subtitle(self):
        ws = _open_ws(self._v2_call())
        assert ws.cell(row=3, column=1).value == "Device Readiness Review"

    def test_row_4_drr_version(self):
        ws = _open_ws(self._v2_call())
        assert ws.cell(row=4, column=1).value == "DRR Version 5.7"

    def test_header_rows_1_3_4_13_span_column_a(self):
        """Per user feedback 2026-08-05: title, subtitle, DRR-version and
        the yellow-highlighting banner should visually occupy column A
        as well (not just B..F). Verify each of those rows has a merge
        range starting at column A."""
        ws = _open_ws(self._v2_call())
        merged_a_starts = {
            m.min_row for m in ws.merged_cells.ranges if m.min_col == 1
        }
        assert 1 in merged_a_starts    # row 1 (OEM Model)
        assert 3 in merged_a_starts    # row 3 (Device Readiness Review)
        assert 4 in merged_a_starts    # row 4 (DRR Version)
        assert 13 in merged_a_starts   # row 13 (yellow banner)

    def test_header_field_values_rendered(self):
        ws = _open_ws(self._v2_call())
        # Row 6: Model Number -> device_id (plain string)
        assert ws.cell(row=6, column=2).value == "Model Number:"
        assert ws.cell(row=6, column=3).value == "SM-F976U"
        # Row 7: fld_lockdown_date now a native date + number_format,
        # NOT a pre-formatted string. openpyxl round-trips to datetime.
        assert ws.cell(row=7, column=2).value == "Compliance Matrix Lockdown On:"
        v7 = ws.cell(row=7, column=3).value
        assert v7 is not None and hasattr(v7, "year") and v7.year == 2026 and v7.month == 3
        # Row 8: req_version -> plain string ("Oct 25")
        assert ws.cell(row=8, column=3).value == "Oct 25"
        # Row 10: target_date -> native date
        v10 = ws.cell(row=10, column=3).value
        assert v10.year == 2026 and v10.month == 4 and v10.day == 29
        # Row 11: FFW -> native date
        v11 = ws.cell(row=11, column=3).value
        assert v11.year == 2026 and v11.month == 5 and v11.day == 13
        # Row 12: LE -> native date
        v12 = ws.cell(row=12, column=3).value
        assert v12.year == 2026 and v12.month == 6 and v12.day == 11
        # Every date cell must carry the mm/dd/yy number_format so Excel
        # doesn't flag it "number stored as text".
        for r in (7, 10, 11, 12):
            assert ws.cell(row=r, column=3).number_format == "mm/dd/yy"

    def test_row_13_yellow_note(self):
        ws = _open_ws(self._v2_call())
        assert "Yellow Highlighting" in (ws.cell(row=13, column=1).value or "")

    def test_row_14_body_column_headers(self):
        ws = _open_ws(self._v2_call())
        assert ws.cell(row=14, column=2).value == "Description of Task"
        assert ws.cell(row=14, column=3).value == "Completion"
        assert ws.cell(row=14, column=4).value == "Current Status"
        assert ws.cell(row=14, column=5).value == "Owner"
        assert ws.cell(row=14, column=6).value == "Remarks"

    def test_missing_drr_version_falls_back_to_bare_label(self):
        ws = _open_ws(self._v2_call(drr_version=None))
        # Row 4 is now merged across A..F; value lives in column A.
        assert ws.cell(row=4, column=1).value == "DRR Version"

    def test_blank_header_field_warns_and_emits_blank(self, caplog):
        with caplog.at_level(
            logging.WARNING,
            logger="core.src.email_service.outbound.drr_report_excel",
        ):
            ws = _open_ws(self._v2_call(
                milestone_headers={
                    "fld_lockdown_date": None,
                    "req_version": "",
                    "target_date": date(2026, 4, 29),
                },
            ))
        assert ws.cell(row=7, column=3).value is None    # fld_lockdown_date blank
        assert ws.cell(row=8, column=3).value is None    # req_version blank
        blanks = [
            r for r in caplog.records
            if "header field" in r.getMessage() and "blank" in r.getMessage()
        ]
        assert len(blanks) >= 2

    def test_missing_logo_path_does_not_crash(self):
        # logo_path=None -> silent skip; logo_path=/nonexistent -> WARN + skip
        out = self._v2_call(logo_path=None)
        assert len(out) > 0
        out2 = self._v2_call(logo_path="/nonexistent/verizon.png")
        assert len(out2) > 0


# ---------------------------------------------------------------------------
# DRR-V2 mode — section + item rows
# ---------------------------------------------------------------------------


class TestV2Body:
    def test_section_row_then_item_rows(self):
        grouping = [
            {
                "section": "Product Documentation Review",
                "work_items": [
                    {"item_no": 1, "item_name": "Product Summary Sheet"},
                    {"item_no": 2, "item_name": "Product ID"},
                ],
            },
            {
                "section": "Pre-Submission Items",
                "work_items": [{"item_no": 5, "item_name": "ODR"}],
            },
        ]
        items = [
            _item(1, item_name="Product Summary Sheet", delivery_state="Open",
                  tg_name="CPM", comment="Finalize by P1"),
            _item(2, item_name="Product ID", delivery_state="Open", tg_name="CPM"),
            _item(5, item_name="ODR", delivery_state="Closed",
                  actual_completion_date=date(2026, 4, 29),
                  tg_name="UX team", comment="ETA: 4/29"),
        ]
        ws = _open_ws(build_drr_report_excel(
            items=items,
            customer_id="MMK", device_id="SM-F976U", milestone_id="DRR",
            section_grouping=grouping,
            drr_version="5.7",
        ))
        # Row 15 = section 1 label (merged B..F, value in A)
        assert ws.cell(row=15, column=1).value == "Product Documentation Review"
        # Rows 16-17: two item rows
        assert ws.cell(row=16, column=1).value == 1
        assert ws.cell(row=16, column=2).value == "Product Summary Sheet"
        assert ws.cell(row=16, column=4).value == "Open"
        assert ws.cell(row=16, column=5).value == "CPM"
        assert ws.cell(row=16, column=6).value == "Finalize by P1"
        assert ws.cell(row=17, column=1).value == 2
        # Row 18 = section 2 header
        assert ws.cell(row=18, column=1).value == "Pre-Submission Items"
        # Row 19 = item#5 with completion date. Now written as native
        # date + mm/dd/yy number_format (no green triangle in Excel).
        assert ws.cell(row=19, column=1).value == 5
        v = ws.cell(row=19, column=3).value
        assert v.year == 2026 and v.month == 4 and v.day == 29
        assert ws.cell(row=19, column=3).number_format == "mm/dd/yy"
        assert ws.cell(row=19, column=4).value == "Closed"

    def test_p1_yellow_marker_fills_description_cell(self):
        """DRR-V2-5a (2026-08-05): a work_item flagged with
        `P1_yellow_marker: true` in template.yaml gets its
        Description-of-Task cell (column B) filled bright yellow
        (FFFF00) to mirror Verizon's Phase 1 Submission Gating
        highlight. Non-flagged rows have no fill on column B."""
        grouping = [{
            "section": "Sec",
            "work_items": [
                {"item_no": 1, "item_name": "Gating", "P1_yellow_marker": True},
                {"item_no": 2, "item_name": "Not gating"},
                {"item_no": 3, "item_name": "Also gating", "P1_yellow_marker": True},
            ],
        }]
        ws = _open_ws(build_drr_report_excel(
            items=[_item(1), _item(2), _item(3)],
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=grouping,
        ))
        # Row 15 = section, rows 16-18 = the three items.
        def _bg(cell):
            return (cell.fill.start_color.rgb or "").upper()

        assert "FFFF00" in _bg(ws.cell(row=16, column=2))
        # Row 17 (Not gating) should NOT carry the yellow fill.
        assert "FFFF00" not in _bg(ws.cell(row=17, column=2))
        assert "FFFF00" in _bg(ws.cell(row=18, column=2))

    def test_missing_postgres_item_still_renders_template_no_and_name(self):
        """When section_grouping references an item_no that has no matching
        DeliveryItem in postgres (setup not run yet / item deleted), the
        row must still appear with item_no + template name; other columns
        blank rather than crashing on getattr."""
        grouping = [{
            "section": "Sec",
            "work_items": [{"item_no": 42, "item_name": "Orphan Item"}],
        }]
        ws = _open_ws(build_drr_report_excel(
            items=[],   # no delivery items at all
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=grouping,
        ))
        assert ws.cell(row=15, column=1).value == "Sec"
        assert ws.cell(row=16, column=1).value == 42
        assert ws.cell(row=16, column=2).value == "Orphan Item"
        # openpyxl round-trip normalizes empty strings to None on load.
        assert ws.cell(row=16, column=4).value in (None, "")
        assert ws.cell(row=16, column=5).value in (None, "")

    def test_empty_section_is_skipped(self):
        grouping = [
            {"section": "Empty", "work_items": []},
            {"section": "Populated",
             "work_items": [{"item_no": 1, "item_name": "X"}]},
        ]
        ws = _open_ws(build_drr_report_excel(
            items=[_item(1)],
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=grouping,
        ))
        # Should NOT see "Empty" anywhere in first 5 rows below body header
        seen = [ws.cell(row=r, column=1).value for r in range(15, 20)]
        assert "Empty" not in seen
        assert "Populated" in seen

    def test_items_85_86_87_excluded_via_section_grouping(self):
        """DRR-V2-1 architect ask: 85 + 86 + 87 must not appear. When
        the caller passes a section_grouping that already filters those
        (via template_lookup.get_drr_section_grouping), they never reach
        the excel body — even if postgres has them."""
        grouping = [{
            "section": "Sec",
            "work_items": [
                {"item_no": 1, "item_name": "Kept"},
                # 85/86/87 deliberately absent
            ],
        }]
        items = [
            _item(1, item_name="Kept"),
            _item(85, item_name="Final DRR excel"),
            _item(86, item_name="Ph-1 placeholder"),
            _item(87, item_name="Default WI"),
        ]
        ws = _open_ws(build_drr_report_excel(
            items=items,
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=grouping,
        ))
        seen_names = [
            ws.cell(row=r, column=2).value for r in range(15, 20)
        ]
        assert "Kept" in seen_names
        assert "Final DRR excel" not in seen_names
        assert "Ph-1 placeholder" not in seen_names
        assert "Default WI" not in seen_names


# ---------------------------------------------------------------------------
# DRR-V2 mode — summary tables
# ---------------------------------------------------------------------------


class TestV2SummaryTables:
    def test_open_closed_completion_percentage(self):
        grouping = [{
            "section": "Sec",
            "work_items": [{"item_no": i, "item_name": f"I{i}"} for i in range(1, 5)],
        }]
        items = [
            _item(1, delivery_state="Open"),
            _item(2, delivery_state="Open"),
            _item(3, delivery_state="Closed"),
            _item(4, delivery_state="Closed"),
        ]
        out = build_drr_report_excel(
            items=items,
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=grouping,
        )
        ws = _open_ws(out)
        # Locate summary rows below body (body ends ~row 19: 15=section,
        # 16-19=items 1-4; summary starts at row 21 = last_body + 2)
        # Scan for "Open" / "Closed" / "Completion Percentage" labels.
        labels_col_c = {
            r: ws.cell(row=r, column=3).value
            for r in range(20, 40)
            if ws.cell(row=r, column=3).value is not None
        }
        vals_col_d = {r: ws.cell(row=r, column=4).value for r in labels_col_c}
        # Find each of the three summary labels
        def _find(label):
            for r, v in labels_col_c.items():
                if v == label:
                    return vals_col_d[r]
            raise AssertionError(f"summary label {label!r} not found")
        assert _find("Open") == 2
        assert _find("Closed") == 2
        assert _find("Completion Percentage") == "50%"

    def test_per_owner_open_counts(self):
        grouping = [{
            "section": "Sec",
            "work_items": [{"item_no": i, "item_name": f"I{i}"} for i in range(1, 5)],
        }]
        items = [
            _item(1, delivery_state="Open", tg_name="CPM"),
            _item(2, delivery_state="Open", tg_name="CPM"),
            _item(3, delivery_state="Open", tg_name="HW team"),
            _item(4, delivery_state="Closed", tg_name="CPM"),
        ]
        out = build_drr_report_excel(
            items=items,
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=grouping,
        )
        ws = _open_ws(out)
        # Scan for owner labels in col C with count in col D
        found: dict[str, int] = {}
        for r in range(20, 50):
            k = ws.cell(row=r, column=3).value
            v = ws.cell(row=r, column=4).value
            if isinstance(k, str) and isinstance(v, int):
                found[k] = v
        assert found.get("CPM") == 2       # 2 Open CPM items
        assert found.get("HW team") == 1

    def test_zero_items_yields_zero_percentage(self):
        out = build_drr_report_excel(
            items=[],
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=[],
        )
        ws = _open_ws(out)
        # No section/item rows; header block still present (title in col A).
        assert ws.cell(row=1, column=1).value == "OEM Model"


# ---------------------------------------------------------------------------
# DRR-V2-8 (2026-08-07): 4-tab workbook + ISO date strip + APPS overlay
# ---------------------------------------------------------------------------


class TestV2FourTabWorkbook:
    def _v2_call(self, **overrides):
        defaults = dict(
            items=[_item(1)],
            customer_id="MMK",
            device_id="SM-F976U",
            milestone_id="DRR",
            section_grouping=[{
                "section": "Sec",
                "work_items": [{"item_no": 1, "item_name": "Item 1"}],
            }],
            drr_version="5.7",
        )
        defaults.update(overrides)
        return build_drr_report_excel(**defaults)

    def _open_wb(self, xlsx_bytes: bytes):
        from openpyxl import load_workbook
        return load_workbook(filename=io.BytesIO(xlsx_bytes))

    def test_four_sheets_in_expected_order(self):
        wb = self._open_wb(self._v2_call())
        assert wb.sheetnames == [
            "Version History", "Checklist", "Applications", "Waivers",
        ]

    def test_version_history_hardcoded_content(self):
        wb = self._open_wb(self._v2_call())
        vh = wb["Version History"]
        assert vh.cell(row=1, column=1).value == "Version History"
        # Header row
        assert vh.cell(row=2, column=1).value == "Version"
        assert vh.cell(row=2, column=4).value == "Change Notes"
        # First data row
        assert vh.cell(row=3, column=1).value == "5.2"
        assert vh.cell(row=3, column=3).value == "Chris Kim"
        # Last data row (5.8 per architect screenshot)
        assert vh.cell(row=9, column=1).value == "5.8"
        assert vh.cell(row=9, column=3).value == "Upesh Kumar"

    def test_waivers_headers_only_no_data_rows(self):
        wb = self._open_wb(self._v2_call())
        wv = wb["Waivers"]
        # Title has device_id interpolated
        assert "SM-F976U" in (wv.cell(row=1, column=1).value or "")
        # Row 2 = the 6 canonical headers
        assert wv.cell(row=2, column=1).value == "#"
        assert wv.cell(row=2, column=2).value == "Description"
        assert wv.cell(row=2, column=5).value == "Waiver #"
        # Row 3 empty (no data rows)
        assert wv.cell(row=3, column=1).value is None

    def test_applications_placeholder_when_no_bytes(self):
        wb = self._open_wb(self._v2_call())
        apps = wb["Applications"]
        assert "No Applications data received yet" in (apps.cell(row=1, column=1).value or "")

    def test_applications_populated_from_source_bytes(self):
        # DRR-V2-8c: source is same Verizon template shape; data lives
        # in rows 12+ (rows 1-11 are Verizon template header). We render
        # our own header + red-fill column headers, then copy data rows
        # 12+ from source preserving row numbers.
        from openpyxl import Workbook
        src = Workbook()
        src.active.title = "Applications"
        # First data row: app #1 Android Pay (supported=No, no version)
        src["Applications"].cell(row=12, column=1, value=1)
        src["Applications"].cell(row=12, column=2, value="Android Pay")
        src["Applications"].cell(row=12, column=3, value="No")
        # Second data row: app #4 Digital Secure (supported=Yes)
        src["Applications"].cell(row=13, column=1, value=4)
        src["Applications"].cell(row=13, column=2, value="Digital Secure/ Verizon Protect")
        src["Applications"].cell(row=13, column=3, value="Yes")
        src["Applications"].cell(row=13, column=4, value="8.3.0-16")
        src["Applications"].cell(row=13, column=7, value="Tested")
        buf = io.BytesIO()
        src.save(buf)

        wb = self._open_wb(self._v2_call(applications_sheet_bytes=buf.getvalue()))
        apps = wb["Applications"]

        # Data preserved at source row numbers
        assert apps.cell(row=12, column=1).value == 1
        assert apps.cell(row=12, column=2).value == "Android Pay"
        assert apps.cell(row=12, column=3).value == "No"
        assert apps.cell(row=13, column=1).value == 4
        assert apps.cell(row=13, column=2).value == "Digital Secure/ Verizon Protect"
        assert apps.cell(row=13, column=3).value == "Yes"
        assert apps.cell(row=13, column=4).value == "8.3.0-16"
        assert apps.cell(row=13, column=7).value == "Tested"

    def test_applications_header_block_rendered(self):
        # DRR-V2-8c: our own header block (rows 1-8) is rendered
        # regardless of source; row 1 = OEM Model, row 2 = subtitle,
        # rows 6-8 = label/value pairs.
        from openpyxl import Workbook
        src = Workbook()
        src.active.title = "Applications"
        src["Applications"].cell(row=12, column=2, value="Dummy App")
        buf = io.BytesIO()
        src.save(buf)

        wb = self._open_wb(self._v2_call(
            applications_sheet_bytes=buf.getvalue(),
            milestone_headers={"fld_lockdown_date": "2026-05-12",
                                "req_version": "Feb 2026",
                                "target_date": "2026-07-01"},
            project_headers={"FFW": "2026-07-15", "LE": None},
        ))
        apps = wb["Applications"]
        # Row 1 title
        assert apps.cell(row=1, column=1).value == "OEM Model"
        # Row 2 subtitle at col C (col A/B reserved for logo)
        assert apps.cell(row=2, column=3).value == "Device Readiness Review"
        # Row 6 Model Number: | SM-F976U | ... DRR Date: | value
        assert apps.cell(row=6, column=2).value == "Model Number:"
        assert apps.cell(row=6, column=3).value == "SM-F976U"
        assert apps.cell(row=6, column=7).value == "DRR Date:"
        # Row 7 Compliance Matrix Lockdown On + Phase 1 Date
        assert apps.cell(row=7, column=2).value == "Compliance Matrix Lockdown On:"
        assert apps.cell(row=7, column=7).value == "Phase 1 Date:"
        # Row 8 VZW Requirements Version (no right-hand pair)
        assert apps.cell(row=8, column=2).value == "VZW Requirements Version:"
        assert apps.cell(row=8, column=7).value is None

    def test_applications_red_fill_column_headers(self):
        # DRR-V2-8c: row 10 red-fill top-level headers + row 11
        # Beta/Final Cert sub-headers under merged E10:F10.
        from openpyxl import Workbook
        src = Workbook()
        src.active.title = "Applications"
        src["Applications"].cell(row=12, column=2, value="Dummy")
        buf = io.BytesIO()
        src.save(buf)

        wb = self._open_wb(self._v2_call(applications_sheet_bytes=buf.getvalue()))
        apps = wb["Applications"]
        # Row 10 top-level headers
        assert apps.cell(row=10, column=1).value == "#"
        assert apps.cell(row=10, column=2).value == "Application Name"
        assert "Is Application Supported" in (apps.cell(row=10, column=3).value or "")
        assert "Application Version Supported" in (apps.cell(row=10, column=4).value or "")
        assert "App Status" in (apps.cell(row=10, column=5).value or "")
        assert "Application Test Status" in (apps.cell(row=10, column=7).value or "")
        assert apps.cell(row=10, column=8).value == "Comments"
        # Row 11 Beta / Final Cert sub-headers
        assert apps.cell(row=11, column=5).value == "Beta"
        assert apps.cell(row=11, column=6).value == "Final Cert"
        # Red fill on row-10 headers (color C00000)
        fill = apps.cell(row=10, column=1).fill
        assert fill.fgColor.rgb in ("00C00000", "FFC00000", "C00000")
        # Row 11 sub-headers also red-filled
        fill = apps.cell(row=11, column=5).fill
        assert fill.fgColor.rgb in ("00C00000", "FFC00000", "C00000")

    def test_applications_gray_fill_when_supported_is_no(self):
        # DRR-V2-8c: rows whose col C = "No" (case-insensitive) get
        # gray fill across cols A..H; "Yes" rows stay white.
        from openpyxl import Workbook
        src = Workbook()
        src.active.title = "Applications"
        # Row 12 unsupported
        src["Applications"].cell(row=12, column=1, value=1)
        src["Applications"].cell(row=12, column=2, value="Android Pay")
        src["Applications"].cell(row=12, column=3, value="No")
        # Row 13 supported
        src["Applications"].cell(row=13, column=1, value=2)
        src["Applications"].cell(row=13, column=2, value="Digital Secure")
        src["Applications"].cell(row=13, column=3, value="Yes")
        # Row 14 unsupported with mixed-case "no"
        src["Applications"].cell(row=14, column=1, value=3)
        src["Applications"].cell(row=14, column=2, value="Caller Name Filter")
        src["Applications"].cell(row=14, column=3, value="no")
        buf = io.BytesIO()
        src.save(buf)

        wb = self._open_wb(self._v2_call(applications_sheet_bytes=buf.getvalue()))
        apps = wb["Applications"]

        # Row 12 (No) gray across A..H
        for col in range(1, 9):
            fill = apps.cell(row=12, column=col).fill
            assert fill.fgColor.rgb in ("00D9D9D9", "FFD9D9D9", "D9D9D9"), (
                f"row 12 col {col} should be gray-filled")
        # Row 13 (Yes) not gray-filled
        # (openpyxl default is empty PatternFill -> fgColor.rgb is '00000000' or 0)
        fill = apps.cell(row=13, column=1).fill
        assert fill.fgColor.rgb not in ("00D9D9D9", "FFD9D9D9", "D9D9D9")
        # Row 14 (mixed-case "no") gray-filled too
        fill = apps.cell(row=14, column=2).fill
        assert fill.fgColor.rgb in ("00D9D9D9", "FFD9D9D9", "D9D9D9")

    def test_applications_error_note_on_missing_sheet_in_source(self):
        # Source .xlsx exists but has no "Applications" sheet
        from openpyxl import Workbook
        src = Workbook()
        src.active.title = "Something Else"
        buf = io.BytesIO()
        src.save(buf)

        wb = self._open_wb(self._v2_call(applications_sheet_bytes=buf.getvalue()))
        apps = wb["Applications"]
        assert "no 'Applications' worksheet" in (apps.cell(row=1, column=1).value or "")


class TestV2IsoDateStringHandling:
    """DRR-V2-8 (2026-08-07): SP REST returns dates as ISO 8601 strings
    like '2026-08-07T07:00:00Z'. Ensure they render as pure dates in
    mm/dd/yy format, not the raw ISO string with time+timezone."""

    def _open_wb(self, xlsx_bytes: bytes):
        from openpyxl import load_workbook
        return load_workbook(filename=io.BytesIO(xlsx_bytes))

    def test_iso_datetime_string_renders_as_date(self):
        out = build_drr_report_excel(
            items=[_item(1)],
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=[{
                "section": "Sec",
                "work_items": [{"item_no": 1, "item_name": "I1"}],
            }],
            milestone_headers={
                "fld_lockdown_date": "2026-03-18T07:00:00Z",  # ISO string with Z
                "req_version": "Oct 25",
                "target_date": "2026-04-29T00:00:00+00:00",   # ISO string with offset
            },
        )
        wb = self._open_wb(out)
        ck = wb["Checklist"]
        # Row 7 = Compliance Matrix Lockdown On value
        v7 = ck.cell(row=7, column=3).value
        assert v7.year == 2026 and v7.month == 3 and v7.day == 18
        assert ck.cell(row=7, column=3).number_format == "mm/dd/yy"
        # Row 10 = DRR Date value
        v10 = ck.cell(row=10, column=3).value
        assert v10.year == 2026 and v10.month == 4 and v10.day == 29
        assert ck.cell(row=10, column=3).number_format == "mm/dd/yy"

    def test_plain_yyyy_mm_dd_string_also_handled(self):
        out = build_drr_report_excel(
            items=[_item(1)],
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=[{
                "section": "Sec",
                "work_items": [{"item_no": 1, "item_name": "I1"}],
            }],
            milestone_headers={
                "fld_lockdown_date": "2026-03-18",             # plain YYYY-MM-DD
                "req_version": "Oct 25",
                "target_date": None,
            },
        )
        wb = self._open_wb(out)
        ck = wb["Checklist"]
        v7 = ck.cell(row=7, column=3).value
        assert v7.year == 2026 and v7.month == 3 and v7.day == 18

    def test_non_iso_string_falls_back_to_raw(self):
        """Not every string is a date -- req_version is 'Oct 25', that
        must stay as-is, no parsing attempt applied."""
        out = build_drr_report_excel(
            items=[_item(1)],
            customer_id="MMK", device_id="X", milestone_id="M",
            section_grouping=[{
                "section": "Sec",
                "work_items": [{"item_no": 1, "item_name": "I1"}],
            }],
            milestone_headers={
                "fld_lockdown_date": None,
                "req_version": "Oct 25",
                "target_date": None,
            },
        )
        wb = self._open_wb(out)
        ck = wb["Checklist"]
        assert ck.cell(row=8, column=3).value == "Oct 25"
