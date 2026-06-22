"""NSD client per [D-013] / [D-041] — two-tree path model + file operations.

IO model (architect alignment 2026-06-11): the corp NSD share is kernel-mounted on
the HILDA host as `hilda-svc` (cifs/smb3, Kerberos keytab) and bind-mounted into the
containers per [D-025] — so routine NSD IO here is plain filesystem IO under the
configured mount root (GlobalStorageConfig.nsd_mount_root; env HILDA_NSD_MOUNT_ROOT).
Async read/write via `aiofiles`; sync stat/readdir wrapped in `asyncio.to_thread`
per the structure-conventions Sync-API wrapping rule. `smbprotocol` is NOT used on
this path — it serves only the diagnostic CLI's share-reachability probe.

Persisted path representation is the share-relative POSIX form (`to_relative()`),
mount-root-independent. UNC rendering (`to_unc()`) is illustrative/diagnostic only.
"""
from __future__ import annotations

import asyncio
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import AsyncIterable, AsyncIterator, Literal

from core.src.diagnostics.error_codes import PipelineError
from core.src.storage.config import get_storage_config

__all__ = [
    "NSDPath",
    "compute_file_hash",
    "extract_first_page",
    "list_inbound_drops",
    "read_file",
    "write_file",
]

_UNC_ROOT = "\\\\share\\hilda"  # illustrative corp share name for diagnostics/docs
_CHUNK = 1 << 16  # 64 KiB


def _mount_root() -> Path:
    return get_storage_config().nsd_mount_root


@dataclass(frozen=True)
class NSDPath:
    """Two-tree NSD path per FR-13. `segments` are relative to the share root."""

    segments: tuple[str, ...]

    # --- Owner inbound tree -------------------------------------------------

    @classmethod
    def inbound_drop(
        cls, customer_id: str, device_id: str, milestone_name: str, item_path_id: str
    ) -> "NSDPath":
        """Owner inbound tree: inbound/<carrier>/<device>/<milestone>/<item>/"""
        return cls(("inbound", customer_id, device_id, milestone_name, item_path_id))

    @classmethod
    def ingress_folder(
        cls, customer_id: str, ingress_nsd: Literal["NSD1", "NSD2"], folder_path: str
    ) -> "NSDPath":
        """FR-77 Type-2 INBOUND folder under the TG's ingress NSD (never outbound target_folder)."""
        parts = tuple(p for p in folder_path.replace("\\", "/").split("/") if p)
        return cls(("inbound", ingress_nsd.lower(), customer_id, *parts))

    # --- HILDA internal tree (FR-86 four path types + zip/outbound zones) ----

    @classmethod
    def internal_classified(
        cls, customer_id, device_id, milestone_name, tg_path_id, item_path_id,
        doc_type, doc_id_slug, rev_number: int,
    ) -> "NSDPath":
        """Classified path: internal/<carrier>/<device>/<milestone>/<tg>/<item>/<doc_type>/<doc_id>/revN/"""
        return cls((
            "internal", customer_id, device_id, milestone_name, tg_path_id,
            item_path_id, doc_type, doc_id_slug, f"rev{rev_number}",
        ))

    @classmethod
    def internal_staged_revision(
        cls, customer_id, device_id, milestone_name, tg_path_id, item_path_id,
        doc_type, original_filename: str,
    ) -> "NSDPath":
        """FR-86 staged-not-revision-determined — aligned but [D-039] ambiguous;
        awaits FR-87 step (C)."""
        return cls((
            "internal", customer_id, device_id, milestone_name, tg_path_id,
            item_path_id, doc_type, "_staged_revision", original_filename,
        ))

    @classmethod
    def internal_staged_classification(
        cls, customer_id, device_id, milestone_name, tg_path_id, item_path_id,
        original_filename: str,
    ) -> "NSDPath":
        """FR-86 staged-not-classified — (item_type, doc_type) misaligned; awaits
        FR-87 step (B). No <doc_type> segment: doc_type is the unresolved dimension."""
        return cls((
            "internal", customer_id, device_id, milestone_name, tg_path_id,
            item_path_id, "_staged_classification", original_filename,
        ))

    @classmethod
    def internal_default_workitem(
        cls, customer_id, device_id, milestone_name, inferred_tg_path_id,
        original_filename: str,
    ) -> "NSDPath":
        """FR-78 + FR-86 + [D-060] unrouted path:
        internal/<carrier>/<device>/<milestone>/<inferred_tg_name>/_unrouted/<filename>
        Pass "_unknown_tg" when inferred_tg_name is NULL (SP-UI-direct-upload edge)."""
        return cls((
            "internal", customer_id, device_id, milestone_name,
            inferred_tg_path_id, "_unrouted", original_filename,
        ))

    @classmethod
    def internal_zip_store(
        cls, customer_id, device_id, milestone_name, tg_path_id, item_path_id,
        original_zip_filename: str,
    ) -> "NSDPath":
        """FR-72 per-item NSD-sourced ZIP storage."""
        return cls((
            "internal", customer_id, device_id, milestone_name, tg_path_id,
            item_path_id, "_zip_store", original_zip_filename,
        ))

    @classmethod
    def internal_unrouted_zip(
        cls, customer_id, device_id, milestone_name, tg_path_id,
        original_zip_filename: str,
    ) -> "NSDPath":
        """FR-72 TG-scoped Email/PLM-sourced ZIP storage."""
        return cls((
            "internal", customer_id, device_id, milestone_name, tg_path_id,
            "_unrouted_zip", original_zip_filename,
        ))

    @classmethod
    def internal_outbound(
        cls, customer_id, device_id, milestone_name, tg_path_id, item_path_id,
        filename: str | None = None,
    ) -> "NSDPath":
        """HILDA-generated artifacts; FR-73 carrier-package zips are transient here."""
        parts = ("internal", customer_id, device_id, milestone_name, tg_path_id,
                 item_path_id, "_outbound")
        return cls(parts + ((filename,) if filename else ()))

    # --- Rendering ------------------------------------------------------------

    def to_relative(self) -> str:
        """Share-relative POSIX path — the persisted representation
        (mount-root-independent; stored on DocumentItemAssociation.local_nsd_path)."""
        return "/".join(self.segments)

    @classmethod
    def from_relative(cls, relative: str) -> "NSDPath":
        """Inverse of to_relative(). Raises STR-E004 on absolute or empty paths."""
        cleaned = relative.replace("\\", "/").strip()
        if not cleaned or cleaned.startswith("/"):
            raise PipelineError(
                "STR-E004", context={"path": relative, "reason": "not a share-relative path"}
            )
        return cls(tuple(p for p in cleaned.split("/") if p))

    def to_local(self) -> Path:
        """Absolute path under the host mount (GlobalStorageConfig.nsd_mount_root)."""
        return _mount_root().joinpath(*self.segments)

    def to_unc(self) -> str:
        """Illustrative UNC rendering of the corp share location — diagnostics and
        human-facing docs only; never the path the code IO sees (see module docstring)."""
        return _UNC_ROOT + "\\" + "\\".join(self.segments)

    @classmethod
    def from_unc(cls, unc: str) -> "NSDPath":
        """Inverse of to_unc() — diagnostic/display use; raises STR-E004 when the
        prefix isn't the HILDA share."""
        if not unc.startswith(_UNC_ROOT + "\\"):
            raise PipelineError("STR-E004", context={"path": unc, "reason": "not under HILDA share"})
        rest = unc[len(_UNC_ROOT) + 1 :]
        return cls(tuple(p for p in rest.split("\\") if p))


# --- File operations (plain IO on the [D-013] host mount) ----------------------


async def read_file(path: NSDPath) -> AsyncIterator[bytes]:
    """Streams file from the NSD mount via aiofiles; used by the hilda-api download
    endpoint (FR-61)."""
    import aiofiles

    local = path.to_local()
    if not local.is_file():
        raise PipelineError(
            "STR-E004", context={"path": path.to_relative(), "reason": "file not found"}
        )
    async with aiofiles.open(local, "rb") as fh:
        while True:
            chunk = await fh.read(_CHUNK)
            if not chunk:
                break
            yield chunk


async def write_file(path: NSDPath, content: AsyncIterable[bytes]) -> None:
    """Writes via the hilda-svc host mount per [D-013]. Idempotent on (path, content) —
    re-write of identical bytes is a no-op. Staged to a temp sibling then atomically
    renamed so readers never see partial files."""
    import aiofiles

    local = path.to_local()
    buf = bytearray()
    async for chunk in content:
        buf.extend(chunk)
    data = bytes(buf)

    try:
        if await asyncio.to_thread(local.is_file):
            existing = await asyncio.to_thread(local.read_bytes)
            if hashlib.sha256(existing).digest() == hashlib.sha256(data).digest():
                return  # idempotent no-op
        await asyncio.to_thread(local.parent.mkdir, *(), **{"parents": True, "exist_ok": True})
        tmp = local.with_name(local.name + ".hilda-partial")
        async with aiofiles.open(tmp, "wb") as fh:
            await fh.write(data)
        await asyncio.to_thread(tmp.replace, local)
    except OSError as exc:
        raise PipelineError(
            "STR-E004", context={"path": path.to_relative(), "reason": str(exc)[:120]}, cause=exc
        )


async def compute_file_hash(path: NSDPath) -> str:
    """SHA-256 per [D-039] Step 0 (exact-duplicate detection)."""
    local = path.to_local()

    def _digest() -> str:
        h = hashlib.sha256()
        try:
            with local.open("rb") as fh:
                for chunk in iter(lambda: fh.read(_CHUNK), b""):
                    h.update(chunk)
        except OSError as exc:
            raise PipelineError(
                "STR-E004", context={"path": path.to_relative(), "reason": str(exc)[:120]}, cause=exc
            )
        return h.hexdigest()

    return await asyncio.to_thread(_digest)


async def list_inbound_drops(
    customer_id: str, device_id: str, milestone_name: str, item_path_id: str
) -> list[NSDPath]:
    """FR-55 polling support — files currently present in the item's inbound folder."""
    base = NSDPath.inbound_drop(customer_id, device_id, milestone_name, item_path_id)
    local = base.to_local()

    def _scan() -> list[NSDPath]:
        if not local.is_dir():
            return []
        return [
            NSDPath(base.segments + (entry.name,))
            for entry in sorted(local.iterdir(), key=lambda e: e.name)
            if entry.is_file()
        ]

    return await asyncio.to_thread(_scan)


async def extract_first_page(path: NSDPath) -> str:
    """First-page text for [D-039] Tier-2 LLM comparison.

    Ph-1 dev scope: plain-text and XLSX (openpyxl, first sheet, first 50 rows).
    PDF / DOCX / DOC extraction awaits the [D-011] profiler library decision
    (pdfplumber vs pypdf vs pymupdf — STATUS.md architecture-phase Next item);
    raises STR-E004 with that pointer until decided.
    """
    local = path.to_local()
    suffix = local.suffix.lower()

    def _extract() -> str:
        if suffix in {".txt", ".csv", ".md", ".log"}:
            return local.read_text(errors="replace")[:4000]
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(local, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            lines = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 50:
                    break
                lines.append("\t".join("" if c is None else str(c) for c in row))
            wb.close()
            return "\n".join(lines)[:4000]
        raise PipelineError(
            "STR-E004",
            context={
                "path": path.to_relative(),
                "reason": f"first-page extraction for '{suffix}' pending [D-011] extraction-library decision",
            },
        )

    return await asyncio.to_thread(_extract)
