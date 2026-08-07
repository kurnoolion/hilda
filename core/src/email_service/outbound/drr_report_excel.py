"""DRR closure final-status Excel report builder.

DRR-V2 shape (2026-08-04) — mirrors the Verizon-issued DRR checklist
template (screenshots from architect 2026-08-04):

  * Rows  1-14  header block:
      row  1  "OEM Model" (merged B..F, centered)
      row  2  Verizon logo (image anchored at A2)
      row  3  "Device Readiness Review" (merged B..F, centered, large)
      row  4  "DRR Version {N}"
      row  5  blank
      row  6  "Model Number:" | <device_id>
      row  7  "Compliance Matrix Lockdown On:" | <fld_lockdown_date>
      row  8  "VZW Requirements Version:" | <req_version>
      row  9  blank
      row 10  "DRR Date:" | <target_date>
      row 11  "Ph1 Date:" | <FFW>
      row 12  "Target TA Date:" | <LE>
      row 13  yellow note "Yellow Highlighting indicates Phase 1
              Submission Gating items"
      row 14  body-column header row (red fill):
              A blank | B Description of Task | C Completion |
              D Current Status | E Owner | F Remarks

  * Row 15+ alternating section rows (bold, gray fill, name spans B..F)
    and item rows (A item_no | B item_name | C actual_completion_date
    | D Open/Closed | E tg_name | F comment). Items 85/86/87 are
    excluded per architect ask (85=Final DRR excel; 86=Ph-1-only
    placeholder; 87=Default WI).

  * Two summary tables after the body:
      - Open / Closed / Completion Percentage
      - Per-owner open-item counts

Header-field mapping (canonical → Verizon label):
    milestone.fld_lockdown_date → Compliance Matrix Lockdown On
    milestone.req_version       → VZW Requirements Version
    milestone.target_date       → DRR Date
    project.FFW                 → Ph1 Date
    project.LE                  → Target TA Date

Legacy flat shape (pre-DRR-V2) still works when the DRR-V2 args
(section_grouping / drr_version / milestone_headers / project_headers
/ logo_path) are all omitted — old callers get the same 4-column
sheet as before.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

__all__ = [
    "build_drr_report_excel",
    "CLOSED_LIKE_STATES",
    "status_for_item",
]

_log = logging.getLogger(__name__)


CLOSED_LIKE_STATES: frozenset[str] = frozenset({
    "Closed",
    "ReadyForSubmission",
    "SubmittedToCustomer",
})


def status_for_item(item: Any) -> str:
    state = getattr(item, "delivery_state", None) or ""
    return "Closed" if state in CLOSED_LIKE_STATES else "Open"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_drr_report_excel(
    items: Iterable[Any] | None = None,
    *,
    customer_id: str | None = None,
    device_id: str | None = None,
    milestone_id: str | None = None,
    section_grouping: list[dict[str, Any]] | None = None,
    drr_version: str | None = None,
    milestone_headers: dict[str, Any] | None = None,
    project_headers: dict[str, Any] | None = None,
    logo_path: str | Path | None = None,
    applications_sheet_bytes: bytes | None = None,
) -> bytes:
    """Return the DRR closure Excel bytes.

    Legacy mode: `build_drr_report_excel(items)` — 4-column flat sheet.
    Preserved so callers that haven't migrated to DRR-V2 still work.

    DRR-V2 mode: all keyword args populated → Verizon-template shape
    with 4 worksheets in this tab order:
      1. Version History  (hardcoded per architect ask 2026-08-07)
      2. Checklist        (main content — header block + section rows +
                           item rows + summary tables)
      3. Applications     (populated from `applications_sheet_bytes` if
                           provided, else a placeholder note)
      4. Waivers          (empty template header row; content pending)

    `applications_sheet_bytes`: raw .xlsx bytes of the APPS TG owner's
    reply attachment. If provided, its "Applications" worksheet is
    copied cell-by-cell into our workbook's Applications tab.
    """
    from openpyxl import Workbook

    materialized = list(items or [])

    v2_mode = section_grouping is not None
    if not v2_mode:
        return _build_legacy_flat(materialized)

    wb = Workbook()
    # Openpyxl auto-creates one default sheet; rename + reuse as Checklist
    # AFTER we've prepended Version History at index 0.
    checklist_ws = wb.active
    checklist_ws.title = "Checklist"

    # Tab 1 — Version History (create at index 0 so it renders first)
    _write_version_history_sheet(wb)

    # Tab 2 — Checklist (main content; already the default sheet, now at index 1)
    _write_header_block(
        ws=checklist_ws,
        customer_id=customer_id or "",
        device_id=device_id or "",
        milestone_id=milestone_id or "",
        drr_version=drr_version,
        milestone_headers=milestone_headers or {},
        project_headers=project_headers or {},
        logo_path=logo_path,
    )
    next_row = _write_body(
        ws=checklist_ws,
        section_grouping=section_grouping or [],
        items=materialized,
    )
    _write_summary_tables(
        ws=checklist_ws,
        start_row=next_row + 2,
        section_grouping=section_grouping or [],
        items=materialized,
    )
    _apply_column_widths(checklist_ws)

    # Tab 3 — Applications (from APPS TG reply attachment, else placeholder)
    _write_applications_sheet(wb, applications_sheet_bytes)

    # Tab 4 — Waivers (empty template header row)
    _write_waivers_sheet(wb, device_id=device_id or "")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet writers — Version History / Applications / Waivers (DRR-V2-8)
# ---------------------------------------------------------------------------


# Static Verizon-provided version-history rows. Never changes per customer
# or per download — hardcoded here so we don't pollute template.yaml with
# what is effectively boilerplate documentation.
_VERSION_HISTORY_ROWS: tuple[tuple[str, str, str, str], ...] = (
    ("5.2", "1/28/2020",  "Chris Kim",     "Added +1 Page confirmation"),
    ("5.3", "1/29/2020",  "Chris Kim",     "Updated Warren IOT to Bedminster IOT"),
    ("5.4", "4/8/2020",   "Chris Kim",     "Updated Sustainability certification form email address"),
    ("5.5", "11/16/2020", "Kevin Cho",     "Added: 5G IOT, 5G lab conformance, and 5G OTA tests for FR1 and FR2"),
    ("5.6", "3/25/2022",  "Kalpesh Kenia", "Added: DACC ID"),
    ("5.7", "2/8/2024",   "Kalpesh Kenia", "Added: eSIM activation for MVA"),
    ("5.8", "2/9/2024",   "Upesh Kumar",   "Added : IT IOT testing"),
)


def _write_version_history_sheet(wb: Any) -> None:
    """Tab 1 — hardcoded Verizon version history table. Same content on
    every download regardless of customer / device / milestone.

    Layout mirrors the architect-provided screenshot:
      row 1: "Version History" (yellow fill, merged A..D)
      row 2: header row "Version | Date | Owner | Change Notes" (peach fill)
      rows 3+: the static rows in _VERSION_HISTORY_ROWS
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title="Version History", index=0)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Row 1 title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws.cell(row=1, column=1, value="Version History")
    c.font = Font(bold=True, size=14)
    c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    c.alignment = center

    # Row 2 header
    hdr_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    for col_idx, label in enumerate(("Version", "Date", "Owner", "Change Notes"), start=1):
        h = ws.cell(row=2, column=col_idx, value=label)
        h.font = Font(bold=True)
        h.fill = hdr_fill
        h.alignment = center

    # Rows 3+ data
    for offset, (ver, dt, owner, notes) in enumerate(_VERSION_HISTORY_ROWS, start=3):
        ws.cell(row=offset, column=1, value=ver).alignment = center
        ws.cell(row=offset, column=2, value=dt).alignment = center
        ws.cell(row=offset, column=3, value=owner).alignment = center
        ws.cell(row=offset, column=4, value=notes).alignment = left

    # Column widths tuned to the screenshot proportions
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 70


def _write_applications_sheet(wb: Any, applications_sheet_bytes: bytes | None) -> None:
    """Tab 3 — Applications data from APPS TG owner's reply attachment.

    When `applications_sheet_bytes` is provided (raw .xlsx bytes of the
    reply), open with openpyxl and copy the "Applications" worksheet
    cell-by-cell into our Applications tab. Column widths are also
    copied so the layout looks like the source.

    When None or empty, render a placeholder note so TPM sees the gap.
    Chunk 2 (2026-08-07) wires the actual APPS-reply retrieval path.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Applications")

    if not applications_sheet_bytes:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        c = ws.cell(
            row=1, column=1,
            value=(
                "No Applications data received yet from APPS TG. "
                "The Applications worksheet from the APPS owner's outreach "
                "reply will render here once that email arrives."
            ),
        )
        c.font = Font(italic=True, color="808080")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions["A"].width = 100
        return

    try:
        from openpyxl import load_workbook
        src_wb = load_workbook(filename=io.BytesIO(applications_sheet_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        _log.warning(
            "drr_report_excel: Applications source .xlsx failed to open: %s: %s",
            type(exc).__name__, str(exc)[:120],
        )
        c = ws.cell(row=1, column=1,
                    value=f"[APPS attachment could not be read: {type(exc).__name__}]")
        c.font = Font(italic=True, color="C00000")
        return

    src = None
    for name in src_wb.sheetnames:
        if name.strip().lower() == "applications":
            src = src_wb[name]
            break
    if src is None:
        _log.warning(
            "drr_report_excel: source APPS xlsx has no 'Applications' sheet "
            "(available: %s)",
            src_wb.sheetnames,
        )
        c = ws.cell(row=1, column=1,
                    value=(
                        "[APPS attachment has no 'Applications' worksheet. "
                        f"Sheets present: {', '.join(src_wb.sheetnames)}]"
                    ))
        c.font = Font(italic=True, color="C00000")
        return

    # Copy cell values (openpyxl cell-copy also lets us pull styles but for
    # Ph-1 we just want the data; a plain value-copy renders readably).
    for row_idx, row in enumerate(src.iter_rows(values_only=True), start=1):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Copy column widths so wide-text columns don't render truncated.
    for col_letter, col_dim in src.column_dimensions.items():
        if col_dim.width:
            ws.column_dimensions[col_letter].width = col_dim.width


def _write_waivers_sheet(wb: Any, *, device_id: str) -> None:
    """Tab 4 — empty Waivers template. Just the title row + column headers;
    content is TPM-authored and will be added in a future chunk.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title="Waivers")

    center = Alignment(horizontal="center", vertical="center")

    # Row 1 title (mirrors "[Tab S12+] Waiver Status" pattern from the
    # architect-provided screenshot; device model interpolated so ops
    # can see which device the waiver context applies to).
    title = f"[{device_id or 'Tab'}] Waiver Status"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=12)
    c.alignment = center

    # Row 2 headers
    hdr_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    headers = ("#", "Description", "Submitted Date", "Approval Date",
               "Waiver #", "VZW / R&D Comments")
    for col_idx, label in enumerate(headers, start=1):
        h = ws.cell(row=2, column=col_idx, value=label)
        h.font = Font(bold=True)
        h.fill = hdr_fill
        h.alignment = center

    # Column widths tuned to the screenshot proportions
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40


# ---------------------------------------------------------------------------
# Header block (rows 1-14)
# ---------------------------------------------------------------------------


_HEADER_ROW_TITLE           = 1
_HEADER_ROW_LOGO            = 2
_HEADER_ROW_SUBTITLE        = 3
_HEADER_ROW_DRR_VERSION     = 4
_HEADER_ROW_MODEL           = 6
_HEADER_ROW_LOCKDOWN        = 7
_HEADER_ROW_REQ_VERSION     = 8
_HEADER_ROW_DRR_DATE        = 10
_HEADER_ROW_PH1_DATE        = 11
_HEADER_ROW_TARGET_TA_DATE  = 12
_HEADER_ROW_YELLOW_NOTE     = 13
_HEADER_ROW_COLUMN_HEADERS  = 14

_HEADER_LABEL_COL = "B"   # right-aligned label
_HEADER_VALUE_COL = "C"   # value

_BODY_HEADERS = (
    "",                     # A: item_no (autofilter, no label)
    "Description of Task",  # B
    "Completion",           # C
    "Current Status",       # D
    "Owner",                # E
    "Remarks",              # F
)


def _write_header_block(
    *,
    ws: Any,
    customer_id: str,
    device_id: str,
    milestone_id: str,
    drr_version: str | None,
    milestone_headers: dict[str, Any],
    project_headers: dict[str, Any],
    logo_path: str | Path | None,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Row 1: "OEM Model" (title, merged A..F, centered)
    ws.merge_cells(start_row=_HEADER_ROW_TITLE, start_column=1,
                   end_row=_HEADER_ROW_TITLE, end_column=6)
    c = ws.cell(row=_HEADER_ROW_TITLE, column=1, value="OEM Model")
    c.font = Font(bold=True, size=16)
    c.alignment = center
    ws.row_dimensions[_HEADER_ROW_TITLE].height = 24

    # Row 2: Verizon logo (image anchored to A2).
    ws.row_dimensions[_HEADER_ROW_LOGO].height = 40
    _embed_logo(ws, logo_path)

    # Row 3: "Device Readiness Review" (merged A..F, centered, large)
    ws.merge_cells(start_row=_HEADER_ROW_SUBTITLE, start_column=1,
                   end_row=_HEADER_ROW_SUBTITLE, end_column=6)
    c = ws.cell(row=_HEADER_ROW_SUBTITLE, column=1, value="Device Readiness Review")
    c.font = Font(bold=True, size=18)
    c.alignment = center
    ws.row_dimensions[_HEADER_ROW_SUBTITLE].height = 28

    # Row 4: "DRR Version {N}" — merged A..F so it visually spans the band
    version_text = f"DRR Version {drr_version}" if drr_version else "DRR Version"
    ws.merge_cells(start_row=_HEADER_ROW_DRR_VERSION, start_column=1,
                   end_row=_HEADER_ROW_DRR_VERSION, end_column=6)
    c = ws.cell(row=_HEADER_ROW_DRR_VERSION, column=1, value=version_text)
    c.font = Font(bold=True, size=11)
    c.alignment = left

    # Rows 6-8, 10-12: label/value pairs. Emit blank + WARN when the raw
    # SP value is None (Q6 answer 2026-08-03). Dates written as native
    # date objects + number_format so Excel doesn't flag them "number
    # stored as text" (green triangle indicator).
    header_rows: list[tuple[int, str, Any, str]] = [
        (_HEADER_ROW_MODEL,           "Model Number:",                  device_id, "device_id"),
        (_HEADER_ROW_LOCKDOWN,        "Compliance Matrix Lockdown On:", milestone_headers.get("fld_lockdown_date"), "fld_lockdown_date"),
        (_HEADER_ROW_REQ_VERSION,     "VZW Requirements Version:",      milestone_headers.get("req_version"), "req_version"),
        (_HEADER_ROW_DRR_DATE,        "DRR Date:",                      milestone_headers.get("target_date"), "target_date"),
        (_HEADER_ROW_PH1_DATE,        "Ph1 Date:",                      project_headers.get("FFW"), "FFW"),
        (_HEADER_ROW_TARGET_TA_DATE,  "Target TA Date:",                project_headers.get("LE"), "LE"),
    ]
    for row, label, raw, field_name in header_rows:
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = Font(bold=True)
        lc.alignment = right
        _write_value_cell(ws, row=row, column=3, raw=raw, alignment=left)
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            _log.warning(
                "drr_report_excel: header field %r blank for "
                "customer=%s device=%s milestone=%s -- cell rendered blank",
                field_name, customer_id, device_id, milestone_id,
            )

    # Row 13: yellow-highlighting note. Full-band A..F, saturated yellow
    # (FFFF00) so it visually pops the way Verizon's template does.
    ws.merge_cells(start_row=_HEADER_ROW_YELLOW_NOTE, start_column=1,
                   end_row=_HEADER_ROW_YELLOW_NOTE, end_column=6)
    c = ws.cell(row=_HEADER_ROW_YELLOW_NOTE, column=1,
                value="Yellow Highlighting indicates Phase 1 Submission Gating items")
    c.font = Font(bold=True, color="9C6500")
    c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    c.alignment = center

    # Row 14: body column headers — red fill.
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, label in enumerate(_BODY_HEADERS, start=1):
        cell = ws.cell(row=_HEADER_ROW_COLUMN_HEADERS, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _embed_logo(ws: Any, logo_path: str | Path | None) -> None:
    """Anchor the verizon.png at cell A2. Silent-skip when path is not
    provided or the file doesn't exist -- corp deploy drops the PNG in
    customizations/branding/; local test runs don't need it."""
    if not logo_path:
        return
    path = Path(logo_path)
    if not path.is_file():
        _log.warning(
            "drr_report_excel: logo path %s not found; skipping image embed", path,
        )
        return
    try:
        from openpyxl.drawing.image import Image as XlImage
        img = XlImage(str(path))
        # Constrain height to ~40px so it fits row 2 without stretching.
        img.height = 40
        img.width = int(img.width * (40 / max(img.height, 1))) if img.height else img.width
        ws.add_image(img, "A2")
    except Exception as exc:  # noqa: BLE001
        _log.warning(
            "drr_report_excel: failed to embed logo %s: %s: %s",
            path, type(exc).__name__, str(exc)[:120],
        )


def _format_header_value(raw: Any) -> str:
    """Kept for backward compat / tests that read the string shape.
    Live cell writes go through _write_value_cell below."""
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.date().strftime("%m/%d/%y")
    if isinstance(raw, date):
        return raw.strftime("%m/%d/%y")
    return str(raw)


# openpyxl date format code — matches "04/29/26" display shown in the
# Verizon template screenshots.
_DATE_NUMBER_FORMAT = "mm/dd/yy"


def _write_value_cell(ws: Any, *, row: int, column: int, raw: Any,
                       alignment: Any) -> None:
    """Write a header value cell with the correct type.

    Dates go in as native `date` objects + a `number_format`, so Excel
    stores them as dates (no green "number stored as text" triangle
    in the upper-left corner) and displays in mm/dd/yy format.

    DRR-V2-8 (2026-08-07): SP REST returns dates as ISO 8601 strings
    like "2026-08-07T07:00:00Z". Parse those to `date` and drop the
    time component instead of dumping the raw ISO string into the
    cell (previous behavior showed the ugly full string).
    """
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        cell = ws.cell(row=row, column=column, value=None)
    elif isinstance(raw, datetime):
        cell = ws.cell(row=row, column=column, value=raw.date())
        cell.number_format = _DATE_NUMBER_FORMAT
    elif isinstance(raw, date):
        cell = ws.cell(row=row, column=column, value=raw)
        cell.number_format = _DATE_NUMBER_FORMAT
    elif isinstance(raw, str):
        # Try ISO 8601 first. datetime.fromisoformat accepts "2026-08-07"
        # AND "2026-08-07T07:00:00" AND "2026-08-07T07:00:00+00:00".
        # SP's "Z" suffix isn't accepted by fromisoformat pre-Py3.11, so
        # normalize to "+00:00" defensively.
        _s = raw.strip()
        _parsed: date | None = None
        try:
            _dt = datetime.fromisoformat(_s.replace("Z", "+00:00"))
            _parsed = _dt.date()
        except ValueError:
            # Not a full ISO datetime. Try plain "YYYY-MM-DD".
            try:
                _parsed = date.fromisoformat(_s[:10])
            except ValueError:
                _parsed = None
        if _parsed is not None:
            cell = ws.cell(row=row, column=column, value=_parsed)
            cell.number_format = _DATE_NUMBER_FORMAT
        else:
            cell = ws.cell(row=row, column=column, value=_s)
    else:
        cell = ws.cell(row=row, column=column, value=str(raw))
    cell.alignment = alignment


# ---------------------------------------------------------------------------
# Body (row 15+): alternating section / item rows
# ---------------------------------------------------------------------------


def _write_body(
    *,
    ws: Any,
    section_grouping: list[dict[str, Any]],
    items: list[Any],
) -> int:
    """Render section rows + item rows below the header block. Returns
    the last row index written (caller uses it to place summary tables).

    Item-row values are drawn from the postgres DeliveryItem for the
    matching item_no; template's work_items provide only the section
    grouping + item ordering.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    section_font = Font(bold=True, size=11)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    # Build an item_no -> DeliveryItem map for fast lookup during body render.
    by_item_no: dict[int, Any] = {}
    for it in items:
        n = getattr(it, "item_no", None)
        if isinstance(n, int):
            by_item_no[n] = it

    current_row = _HEADER_ROW_COLUMN_HEADERS + 1  # row 15

    for group in section_grouping:
        section_name = group.get("section")
        work_items = group.get("work_items") or []
        if not work_items:
            continue

        # Section row: bold gray band across A..F, name in B.
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=6)
        sc = ws.cell(row=current_row, column=1,
                     value=str(section_name) if section_name else "(unnamed section)")
        sc.font = section_font
        sc.fill = section_fill
        sc.alignment = left
        current_row += 1

        for wi in work_items:
            wi_no = wi.get("item_no")
            wi_name = wi.get("item_name") or ""
            wi_p1_gating = bool(wi.get("P1_yellow_marker"))
            di = by_item_no.get(wi_no) if isinstance(wi_no, int) else None
            _write_item_row(
                ws=ws,
                row=current_row,
                item_no=wi_no,
                fallback_item_name=wi_name,
                delivery_item=di,
                p1_gating=wi_p1_gating,
                left=left, center=center,
            )
            current_row += 1

    return current_row - 1


# Bright yellow used on the row-13 banner AND on P1 gating item
# Description-of-Task cells so the two visually agree.
_P1_YELLOW_HEX = "FFFF00"


def _write_item_row(
    *,
    ws: Any,
    row: int,
    item_no: Any,
    fallback_item_name: str,
    delivery_item: Any,
    p1_gating: bool,
    left: Any,
    center: Any,
) -> None:
    """One data row: A item_no | B name | C completion | D status |
    E owner | F remarks. When p1_gating=True, the Description-of-Task
    cell (B) is filled bright yellow to mirror Verizon's Phase 1
    Submission Gating highlight."""
    from openpyxl.styles import PatternFill

    if delivery_item is not None:
        item_name = (
            getattr(delivery_item, "item_name", None)
            or fallback_item_name
            or f"Item {item_no or '?'}"
        )
        completion_raw = getattr(delivery_item, "actual_completion_date", None)
        status = status_for_item(delivery_item)
        owner = getattr(delivery_item, "tg_name", None) or ""
        remarks = getattr(delivery_item, "comment", None) or ""
    else:
        # No postgres DeliveryItem for this template work_item (setup
        # hasn't run yet, or the item was deleted). Render item_no +
        # template name only; leave the rest blank.
        item_name = fallback_item_name or f"Item {item_no or '?'}"
        completion_raw = None
        status = ""
        owner = ""
        remarks = ""

    a = ws.cell(row=row, column=1, value=item_no)
    a.alignment = center
    b = ws.cell(row=row, column=2, value=str(item_name))
    b.alignment = left
    if p1_gating:
        b.fill = PatternFill(start_color=_P1_YELLOW_HEX,
                             end_color=_P1_YELLOW_HEX,
                             fill_type="solid")
    # Completion date: write native date + number_format so no green
    # "number stored as text" marker appears.
    _write_value_cell(ws, row=row, column=3, raw=completion_raw, alignment=center)
    d = ws.cell(row=row, column=4, value=status)
    d.alignment = center
    e = ws.cell(row=row, column=5, value=str(owner))
    e.alignment = left
    f = ws.cell(row=row, column=6, value=str(remarks))
    f.alignment = left


# ---------------------------------------------------------------------------
# Summary tables (bottom of sheet)
# ---------------------------------------------------------------------------


def _write_summary_tables(
    *,
    ws: Any,
    start_row: int,
    section_grouping: list[dict[str, Any]],
    items: list[Any],
) -> None:
    """Two side-by-side / stacked mini-tables:
      1. Open / Closed / Completion Percentage
      2. Per-owner (tg_name) open-item count
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    label_font = Font(bold=True)
    fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Filter items to the ones surfaced by section_grouping (already
    # excludes item#85/86/87) so counts don't double-count excluded ones.
    surfaced_item_nos: set[int] = set()
    for group in section_grouping:
        for wi in group.get("work_items") or []:
            n = wi.get("item_no")
            if isinstance(n, int):
                surfaced_item_nos.add(n)
    surfaced_items = [
        it for it in items
        if isinstance(getattr(it, "item_no", None), int)
        and it.item_no in surfaced_item_nos
    ]

    open_count = sum(1 for it in surfaced_items if status_for_item(it) == "Open")
    closed_count = sum(1 for it in surfaced_items if status_for_item(it) == "Closed")
    total = open_count + closed_count
    pct = (closed_count / total * 100.0) if total else 0.0

    rows: list[tuple[str, Any]] = [
        ("Open", open_count),
        ("Closed", closed_count),
        ("Completion Percentage", f"{pct:.0f}%"),
    ]
    for offset, (label, value) in enumerate(rows):
        r = start_row + offset
        lc = ws.cell(row=r, column=3, value=label)
        lc.font = label_font
        lc.alignment = center
        vc = ws.cell(row=r, column=4, value=value)
        vc.alignment = center
        if label == "Completion Percentage":
            lc.fill = fill
            vc.fill = fill

    # Per-owner open counts.
    owner_open: dict[str, int] = {}
    for it in surfaced_items:
        if status_for_item(it) == "Open":
            owner = (getattr(it, "tg_name", None) or "").strip() or "(unassigned)"
            owner_open[owner] = owner_open.get(owner, 0) + 1
    if owner_open:
        header_row = start_row + len(rows) + 2
        for offset, (owner, count) in enumerate(
            sorted(owner_open.items(), key=lambda kv: kv[0])
        ):
            r = header_row + offset
            lc = ws.cell(row=r, column=3, value=owner)
            lc.font = label_font
            lc.alignment = center
            lc.fill = fill
            vc = ws.cell(row=r, column=4, value=count)
            vc.alignment = center


# ---------------------------------------------------------------------------
# Column widths (shared by legacy + V2)
# ---------------------------------------------------------------------------


def _apply_column_widths(ws: Any) -> None:
    ws.column_dimensions["A"].width = 8    # Item No
    ws.column_dimensions["B"].width = 60   # Description of Task
    ws.column_dimensions["C"].width = 14   # Completion (date)
    ws.column_dimensions["D"].width = 16   # Current Status
    ws.column_dimensions["E"].width = 14   # Owner
    ws.column_dimensions["F"].width = 30   # Remarks


# ---------------------------------------------------------------------------
# Legacy 4-column flat sheet (pre-DRR-V2)
# ---------------------------------------------------------------------------


def _build_legacy_flat(materialized: list[Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "DRR Closure"

    headers = ("Item No", "Item Title", "Open/Closed Status", "Owner Comment")
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    materialized.sort(key=lambda it: (getattr(it, "item_no", 0) or 0))
    for row_offset, it in enumerate(materialized, start=2):
        item_no = getattr(it, "item_no", None)
        item_name = (
            getattr(it, "item_name", None)
            or getattr(it, "item_title", None)
            or f"Item {item_no or '?'}"
        )
        status = status_for_item(it)
        note = getattr(it, "owner_status_note", None) or ""
        ws.cell(row=row_offset, column=1, value=item_no)
        ws.cell(row=row_offset, column=2, value=str(item_name))
        ws.cell(row=row_offset, column=3, value=status)
        ws.cell(row=row_offset, column=4, value=str(note))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
