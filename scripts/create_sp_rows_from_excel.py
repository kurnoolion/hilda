"""create_sp_rows_from_excel.py -- bulk-create SharePoint list rows from an Excel sheet.

Standalone helper for onboarding a new milestone template. Read-then-write:
opens an .xlsx file, treats row 1 as the header row (internal SP column
names, in any order), and treats each subsequent non-empty row as one SP
item to create in the target list.

USE CASE (2026-07-15 architect ask): populate Deliverables_Template with
~85 rows for the DRR milestone (vs the P1 test milestone's 10 items),
without hand-clicking in the SP UI.

Read row 1 -> {col_name: col_index}. For each data row: build
{col_name: value} using only the cells whose column appears in row 1,
skip empty rows, then POST via `SpSession.create(list_name, "", fields)`.

SAFETY RAILS:
  * By default, refuses to write to a list whose name doesn't contain
    "Template" -- prevents accidental writes to Deliverables_MMK etc.
    Override with --force.
  * --dry-run prints what would be created without hitting SP.
  * Excel cells are coerced conservatively: empty -> skipped from payload
    (SP uses column default), booleans as-is, integers/floats as-is,
    datetimes -> ISO string, strings stripped. Callers can pre-encode
    complex fields (Choice multi-value, list-of-list) as strings in the
    Excel cell -- this script does not attempt to guess their shape.
  * Per-row error handling: one row failing does not abort the batch.
    Failed rows are logged with the row number and the first ~200 chars
    of the SP error body; the run continues.

USAGE (inside the hilda-worker container):

    # Copy in and run
    podman cp scripts/create_sp_rows_from_excel.py hilda-worker:/tmp/
    podman cp <your_excel>.xlsx hilda-worker:/tmp/rows.xlsx

    # Dry-run first (no writes)
    podman exec -e PYTHONPATH=/app hilda-worker python /tmp/create_sp_rows_from_excel.py \\
        /tmp/rows.xlsx Deliverables_Template --dry-run

    # If the dry-run looks correct, do it for real:
    podman exec -e PYTHONPATH=/app hilda-worker python /tmp/create_sp_rows_from_excel.py \\
        /tmp/rows.xlsx Deliverables_Template

    # Optional flags:
    #   --sheet <name>       select a non-default worksheet
    #   --start-row N        skip first N data rows (useful for retry-from-failure)
    #   --limit N            create only the first N data rows
    #   --force              allow writing to a non-*Template* list name
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

# Ensure the container's /app root is on sys.path even when python is invoked
# from /tmp (mirrors list_sp_columns.py fix).
for _candidate in ("/app", str(Path(__file__).resolve().parents[1])):
    if Path(_candidate).is_dir() and _candidate not in sys.path:
        sys.path.insert(0, _candidate)


def _resolve_config_path() -> Path | None:
    for candidate in (
        Path("config/sharepoint_integration.json"),
        Path("/app/config/sharepoint_integration.json"),
    ):
        if candidate.exists():
            return candidate
    return None


def _coerce_cell(value: Any) -> Any | None:
    """Convert Excel cell value to what SP REST expects. Returns None to
    indicate 'skip this field' (empty cell -> let SP apply its default)."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    # Fallback: stringify + strip. Prevents unexpected openpyxl-specific
    # objects from confusing SP.
    text = str(value).strip()
    return text if text else None


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Bulk-create SharePoint list rows from an Excel sheet."
    )
    parser.add_argument("xlsx_path", help="Path to .xlsx file with row 1 = column names.")
    parser.add_argument("list_name", help="Target SP list (e.g., Deliverables_Template).")
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print payloads without hitting SP.")
    parser.add_argument("--start-row", type=int, default=0,
                        help="Skip the first N data rows (0-indexed relative to data rows; "
                             "useful for retrying a partial batch).")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only the first N data rows after --start-row.")
    parser.add_argument("--force", action="store_true",
                        help="Allow writing to a list whose name doesn't contain 'Template'. "
                             "Without this flag the script refuses non-template lists as a "
                             "safety rail.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv if argv is not None else sys.argv[1:])

    xlsx = Path(args.xlsx_path)
    if not xlsx.is_file():
        print(f"ERROR: not a file: {xlsx}")
        return 2

    # Safety rail
    if "template" not in args.list_name.lower() and not args.force:
        print(f"ERROR: list name '{args.list_name}' does not contain 'Template'. Refusing "
              f"to write to a non-template list without --force. Rerun with --force if "
              f"you're sure.")
        return 3

    # Config + session
    try:
        from core.src.sharepoint_integration.config import GlobalSharePointConfig
        from core.src.sharepoint_integration.sp_session import SpSession
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: HILDA import failed: {type(exc).__name__}: {exc}")
        print("Run inside the hilda-worker container with PYTHONPATH=/app.")
        return 4

    cfg = GlobalSharePointConfig.from_sources(config_path=_resolve_config_path())
    if not cfg.username or not cfg.password:
        print("ERROR: NTLM credentials missing. Set them in "
              "config/sharepoint_integration.json or via HILDA_SP_* env vars.")
        return 5

    # openpyxl available in the container (declared in requirements.txt)
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run this inside the hilda-worker container.")
        return 6

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        print("ERROR: empty worksheet -- no header row.")
        return 7

    # Column index -> header name. Skip cells whose header is blank so a
    # trailing empty column in Excel doesn't create phantom fields.
    header_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            header_map[idx] = name
    if not header_map:
        print("ERROR: header row had no non-empty cells.")
        return 8

    print(f"[info] site_url : {cfg.site_url}")
    print(f"[info] username : {cfg.username}")
    print(f"[info] list_name: {args.list_name}")
    print(f"[info] xlsx     : {xlsx}")
    print(f"[info] sheet    : {ws.title}")
    print(f"[info] columns  : {sorted(header_map.values())}")
    print(f"[info] mode     : {'DRY-RUN (no writes)' if args.dry_run else 'LIVE (SP writes)'}")
    print()

    # Materialize data rows so we can report totals + apply --start-row/--limit.
    data_rows: list[tuple[int, tuple[Any, ...]]] = []
    for excel_row_num, row in enumerate(rows_iter, start=2):  # excel row 2 = first data row
        if all(cell is None for cell in row):
            continue  # skip fully-empty rows silently
        data_rows.append((excel_row_num, row))

    total_data_rows = len(data_rows)
    if args.start_row > 0:
        data_rows = data_rows[args.start_row:]
    if args.limit is not None:
        data_rows = data_rows[: args.limit]

    print(f"[info] total data rows in sheet : {total_data_rows}")
    print(f"[info] rows to process this run : {len(data_rows)}"
          + (f" (start-row={args.start_row})" if args.start_row else "")
          + (f" (limit={args.limit})" if args.limit is not None else ""))
    print()

    # Session lazily. Skip session build entirely in dry-run.
    session = None
    if not args.dry_run:
        session = SpSession(
            site_url=cfg.site_url,
            ntlm_user=cfg.username,
            ntlm_pass=cfg.password,
        )

    created = 0
    failed = 0
    skipped_empty = 0
    failed_rows: list[tuple[int, str]] = []

    for i, (excel_row_num, row) in enumerate(data_rows, start=1):
        fields: dict[str, Any] = {}
        for col_idx, col_name in header_map.items():
            cell_value = row[col_idx] if col_idx < len(row) else None
            coerced = _coerce_cell(cell_value)
            if coerced is not None:
                fields[col_name] = coerced

        if not fields:
            skipped_empty += 1
            continue

        if args.dry_run:
            print(f"[dry-run] excel-row={excel_row_num}: would create "
                  f"{json.dumps(fields, default=str, sort_keys=True)}")
            created += 1
        else:
            try:
                status, body = session.create(
                    list_name=args.list_name,
                    customer_id="",
                    fields=fields,
                )
                if 200 <= status < 300:
                    created += 1
                else:
                    failed += 1
                    err = _short_error(body)
                    failed_rows.append((excel_row_num, f"HTTP {status}: {err}"))
                    print(f"[fail] excel-row={excel_row_num} HTTP {status}: {err}")
            except Exception as exc:  # noqa: BLE001
                failed += 1
                msg = f"{type(exc).__name__}: {str(exc)[:180]}"
                failed_rows.append((excel_row_num, msg))
                print(f"[fail] excel-row={excel_row_num} {msg}")

        # Progress every 10 rows in live mode
        if not args.dry_run and i % 10 == 0:
            print(f"[progress] processed {i}/{len(data_rows)} "
                  f"(created={created} failed={failed})")

    print()
    print(f"[done] created={created} failed={failed} "
          f"skipped_empty={skipped_empty} rows_processed={len(data_rows)}")
    if failed_rows:
        print("\nFAILED ROWS (excel-row-number, error):")
        for row_num, err in failed_rows:
            print(f"  row {row_num}: {err}")
        print("\nRe-run with --start-row set past the last successful row to retry the "
              "tail without re-creating already-created rows.")
        return 1
    return 0


def _short_error(body: Any) -> str:
    """Extract a compact error message from SP's JSON response."""
    if isinstance(body, dict):
        # SP odata=verbose typically nests errors under body["error"]["message"]["value"]
        err = body.get("error") or {}
        msg = err.get("message") if isinstance(err, dict) else None
        if isinstance(msg, dict):
            val = msg.get("value")
            if val:
                return str(val)[:200]
        if err:
            return str(err)[:200]
    return str(body)[:200] if body else ""


if __name__ == "__main__":
    raise SystemExit(main())
