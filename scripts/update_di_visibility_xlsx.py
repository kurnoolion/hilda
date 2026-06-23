"""Update docs/sp_ui_engineer/DeliveryItem_visibility_review.xlsx for per_item_rule_overrides Ph-2 deferral.

Per user 2026-06-12: move per_item_rule_overrides from must_show / yes to
not_to_be_shown / no with Ph-2 deferred note. TPM rule overrides for Ph-1
go through ops YAML drop-zone (Option a from 2026-06-12 review).
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

PATH = "docs/sp_ui_engineer/DeliveryItem_visibility_review.xlsx"

COLOR_NOT_SHOWN = "F4B084"  # orange
COLOR_NO = "F4B084"

wb = load_workbook(PATH)
s = wb["DI field visibility"]

for r in range(2, s.max_row + 1):
    if s.cell(row=r, column=1).value == "per_item_rule_overrides":
        s.cell(row=r, column=2).value = "not_to_be_shown"
        s.cell(row=r, column=2).fill = PatternFill("solid", start_color=COLOR_NOT_SHOWN)
        s.cell(row=r, column=3).value = "no"
        s.cell(row=r, column=3).fill = PatternFill("solid", start_color=COLOR_NO)
        s.cell(row=r, column=4).value = (
            "Ph-2 DEFERRED per 2026-06-12 SP UI engineer review — TPM rule overrides "
            "for Ph-1 handled via ops YAML drop-zone (customizations/rules/<customer_slug>/per_item_overrides.yaml), "
            "not via SP DeliveryItems list. Revisit in Ph-2 when concrete TPM-facing overrides crystallize "
            "(recommended Ph-2 high-value subset: reminder_interval_override_days, escalation_threshold_override_days, "
            "stale_alert_suppressed, all_rules_paused — 4 discrete SP columns instead of one abstract Note field)."
        )
        break

wb.save(PATH)
print(f"Updated: {PATH}")
