"""Add customer_slug + device_slug to Milestones tab in HILDA_SP_Schema.xlsx per D-DRAFT-Z 2026-06-12.

These are denormalized columns on Milestone rows so HILDA can derive (customer, device)
context without reading Customers + Devices SP lists at runtime.
"""
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

PATH = "docs/sp_ui_engineer/HILDA_SP_Schema.xlsx"


def style_from(src_cell):
    return {
        "font": copy(src_cell.font),
        "fill": copy(src_cell.fill),
        "alignment": copy(src_cell.alignment),
        "border": copy(src_cell.border),
        "number_format": src_cell.number_format,
    }


def apply_style(cell, sty):
    cell.font = sty["font"]
    cell.fill = sty["fill"]
    cell.alignment = sty["alignment"]
    cell.border = sty["border"]
    cell.number_format = sty["number_format"]


def write_row(sheet, row, values, style_row=4):
    for col_idx, val in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=col_idx)
        src = sheet.cell(row=style_row, column=col_idx)
        sty = style_from(src)
        cell.value = val
        apply_style(cell, sty)


wb = load_workbook(PATH, data_only=False)
s = wb["Milestones"]

# Append 2 new rows
next_row = s.max_row + 1

write_row(s, next_row, [
    "customer_slug",
    "customer_slug",
    "customer_slug",
    "Text",
    "",
    "Y",
    "Ph-1",
    "",
    "D-DRAFT-Z 2026-06-12 (HILDA runtime SP coupling restricted to Milestones + DeliveryItems lists)",
    "READ-ONLY mirror per D-DRAFT-Z; denormalized from customizations/template_schemas/<customer_slug>/customer.yaml at HILDA startup. HILDA reads this from Milestone row at runtime instead of joining Customers SP list. TPM SP UI must not allow editing."
], style_row=4)
next_row += 1

write_row(s, next_row, [
    "device_slug",
    "device_slug",
    "device_slug",
    "Text",
    "",
    "Y",
    "Ph-1",
    "",
    "D-DRAFT-Z 2026-06-12",
    "READ-ONLY mirror per D-DRAFT-Z; denormalized from customer.yaml devices: sub-block at HILDA startup. HILDA reads from Milestone row at runtime instead of joining Devices SP list. TPM SP UI must not allow editing."
], style_row=4)

# Also annotate Customers/Devices/Users/PMCredentials tabs as SP-display-only
def annotate_sp_display_only(sheet_name, note):
    s2 = wb[sheet_name]
    # Update row 2 (description row) to add D-DRAFT-Z note
    cur = s2.cell(row=2, column=1).value or ""
    if "D-DRAFT-Z" not in cur:
        s2.cell(row=2, column=1).value = f"{cur} | D-DRAFT-Z 2026-06-12: SP UI engineer display surface only; HILDA does NOT read this list at runtime — {note}"

annotate_sp_display_only("Customers", "customer data sourced from customizations/template_schemas/<customer_slug>/customer.yaml at HILDA startup")
annotate_sp_display_only("Devices", "device data sourced from customer.yaml devices: sub-block")
annotate_sp_display_only("Users", "owner identity denormalized onto DeliveryItem rows per FR-2 owner identity model 2026-06-12; HILDA reads from DI rows, not Users list")
annotate_sp_display_only("PMCredentials", "credentials live in HILDA-local credential_service per [D-019] / [D-038]; HILDA reads via get_credential() not via SP")

wb.save(PATH)
print(f"Saved {PATH}")
print(f"Milestones rows now: {wb['Milestones'].max_row}")
