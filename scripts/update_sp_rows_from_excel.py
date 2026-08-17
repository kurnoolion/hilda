"""update_sp_rows_from_excel.py -- bulk-update SharePoint list rows from an Excel sheet.

Companion to create_sp_rows_from_excel.py. That script BLINDLY CREATES rows;
this one READS existing rows, MATCHES by a key column (default: item_no),
and MERGE-updates only the columns you specify. No duplicate rows will
be created -- if a key value in the Excel sheet has no matching SP row,
it's reported and skipped.

USE CASE (2026-08-16 architect ask): update tg_name + owner_name +
owner_corp_email + owner_corp_id on existing DRR rows in
Deliverables_Template (the SP UI engineer copies from this template to
Deliverables_MMK when the TPM clicks "setup milestone/deliverables";
updating the template so future setups get the new owners).

HOW IT WORKS:
  1. Read Excel row 1 as header row (SP internal column names, any order).
  2. Read the target SP list via GET (with an optional --filter to scope by
     milestone_id / project_model / etc.). Build item_no -> sp_id map.
  3. For each Excel data row:
       * find the SP row whose key column value matches
       * for each allowed field (--fields), compare Excel vs SP; if the
         value differs (or --always-send is on), include it in the MERGE
         payload
       * if the payload is non-empty AND not --dry-run, POST MERGE.
  4. Report totals: updated / no-diff / unmatched-key / duplicate-key /
     failed.

FIELD ALLOWLIST (default, per architect 2026-08-16):
  --fields tg_name,owner_name,owner_corp_email,owner_corp_id

Every other Excel column is IGNORED. Pass --fields explicitly to widen /
narrow the set. Owner fields are semicolon-separated multi-value text
per architect multi-owner direction 2026-08-14 (HILDA parses via
_split_owner_list at ingest).

SAFETY RAILS:
  * By default, refuses to write to a list whose name doesn't contain
    "Template" -- prevents accidental writes to Deliverables_MMK.
    Override with --force.
  * --dry-run prints what would be MERGEd without hitting SP.
  * Per-row error handling: one row failing does not abort the batch.
  * Duplicate-key detection: if two SP rows share the same key value
    (e.g. two milestones' item_no=5 in an unfiltered read), those rows
    are SKIPPED with a warning -- pass --filter to narrow the scope.

USAGE (inside the hilda-worker container):

    podman cp scripts/update_sp_rows_from_excel.py hilda-worker:/tmp/
    podman cp <your_excel>.xlsx hilda-worker:/tmp/rows.xlsx

    # Dry-run first
    podman exec -e PYTHONPATH=/app hilda-worker python /tmp/update_sp_rows_from_excel.py \\
        /tmp/rows.xlsx Deliverables_Template \\
        --filter milestone_id=DRR \\
        --dry-run

    # Live run
    podman exec -e PYTHONPATH=/app hilda-worker python /tmp/update_sp_rows_from_excel.py \\
        /tmp/rows.xlsx Deliverables_Template \\
        --filter milestone_id=DRR

    # Optional flags:
    #   --key-col NAME        Excel/SP column used as the natural key (default: item_no)
    #   --fields A,B,C        override the field allowlist (default: 4 owner fields)
    #   --filter col=val      OData $filter clause to narrow the SP read (recommended)
    #   --sheet NAME          worksheet name (default: first sheet)
    #   --always-send         send every allowed field even when unchanged (skip diff)
    #   --limit N             process only first N matched Excel rows
    #   --force               allow writing to a non-*Template* list
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

for _candidate in ("/app", str(Path(__file__).resolve().parents[1])):
    if Path(_candidate).is_dir() and _candidate not in sys.path:
        sys.path.insert(0, _candidate)


_DEFAULT_FIELDS = ("tg_name", "owner_name", "owner_corp_email", "owner_corp_id")
_DEFAULT_KEY_COL = "item_no"


def _resolve_config_path() -> Path | None:
    for candidate in (
        Path("config/sharepoint_integration.json"),
        Path("/app/config/sharepoint_integration.json"),
    ):
        if candidate.exists():
            return candidate
    return None


def _coerce_cell(value: Any) -> Any | None:
    """Excel cell -> SP-REST-compatible value. None = 'skip'."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else None


def _norm_for_compare(v: Any) -> str:
    """Normalize a value for diff comparison. Owner fields are text with
    semicolons + varying whitespace; treat 'alice@corp; bob@corp' and
    'alice@corp;bob@corp' as equal so we don't spam SP with cosmetic
    round-trips."""
    if v is None:
        return ""
    s = str(v).strip()
    # collapse runs of whitespace, strip whitespace around each ';' segment
    parts = [p.strip() for p in s.split(";")]
    parts = [p for p in parts if p]
    return "; ".join(parts).lower() if parts else s.lower()


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Bulk-update SharePoint list rows from an Excel sheet "
                    "(key-based match; MERGE-updates specified columns only)."
    )
    parser.add_argument("xlsx_path", help="Path to .xlsx file with row 1 = SP column names.")
    parser.add_argument("list_name", help="Target SP list (default-safe: Deliverables_Template).")
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet).")
    parser.add_argument("--key-col", default=_DEFAULT_KEY_COL,
                        help=f"Column name used to match Excel row <-> SP row "
                             f"(default: {_DEFAULT_KEY_COL}).")
    parser.add_argument("--fields", default=",".join(_DEFAULT_FIELDS),
                        help=f"Comma-separated allowlist of fields to update. "
                             f"Every other Excel column is ignored. "
                             f"Default: {','.join(_DEFAULT_FIELDS)}")
    parser.add_argument("--filter", dest="odata_filter", default=None,
                        help="OData $filter clause to narrow the SP-side read, "
                             "as 'col=val' (translated to \"col eq 'val'\") OR "
                             "a full OData expression passed through. "
                             "STRONGLY RECOMMENDED when the list holds multiple "
                             "milestones (else duplicate item_no across milestones "
                             "will cause ambiguous-key skips). Example: "
                             "--filter milestone_id=DRR")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print MERGE payloads without hitting SP.")
    parser.add_argument("--always-send", action="store_true",
                        help="Send every allowed field even if the Excel value "
                             "matches SP. Default is diff-only (skip unchanged).")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only the first N matched Excel rows.")
    parser.add_argument("--force", action="store_true",
                        help="Allow writing to a non-*Template* list.")
    return parser.parse_args(argv)


def _translate_filter(raw: str) -> str:
    """Turn a shorthand 'col=val' into an OData $filter clause. If raw
    already contains 'eq'/'ne'/'and'/'or', pass it through unchanged."""
    lowered = raw.lower()
    for op in (" eq ", " ne ", " and ", " or ", " gt ", " lt ", " ge ", " le "):
        if op in lowered:
            return raw
    if "=" in raw:
        col, val = raw.split("=", 1)
        col = col.strip()
        val = val.strip().strip("'\"")
        # Quote as string; if numeric-looking, SP accepts unquoted too.
        return f"{col} eq '{val}'"
    return raw


def _fetch_all_sp_rows(
    session, list_name: str, key_col: str, extra_cols: list[str],
    odata_filter: str | None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Page through the SP list; return (rows, warnings). Each row is a
    dict with '_sp_id' plus the requested $select columns.
    """
    select_cols = ["Id", key_col, *extra_cols]
    select_str = ",".join(select_cols)
    top = 500  # SP page size

    url = f"/_api/web/lists/getbytitle('{list_name}')/items"
    params: dict[str, str] = {
        "$select": select_str,
        "$top":    str(top),
    }
    if odata_filter:
        params["$filter"] = _translate_filter(odata_filter)

    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    next_url: str | None = None
    first = True
    while True:
        resp = session.get(next_url or url, params=(params if first else None))
        first = False
        if resp.status_code != 200:
            body = resp.text[:200] if resp.text else ""
            raise RuntimeError(f"SP GET failed ({resp.status_code}): {body}")
        payload = resp.json() if resp.content else {}
        # SP odata=verbose nests under d.results; odata=nometadata surfaces value.
        inner = payload.get("d") if isinstance(payload, dict) else None
        if isinstance(inner, dict):
            results = inner.get("results") or []
            next_link = inner.get("__next")
        else:
            results = payload.get("value") or []
            next_link = payload.get("odata.nextLink") or payload.get("@odata.nextLink")
        for item in results:
            row: dict[str, Any] = {"_sp_id": item.get("Id") or item.get("ID")}
            for col in [key_col, *extra_cols]:
                row[col] = item.get(col)
            rows.append(row)
        if next_link:
            next_url = next_link
        else:
            break
    return rows, warnings


def _index_by_key(
    rows: list[dict[str, Any]], key_col: str,
) -> tuple[dict[str, dict[str, Any]], set[str]]:
    """Build {key_value_normalized: sp_row}. Returns (index, duplicate_keys).
    Duplicate keys are excluded from the index entirely so we don't
    silently pick one."""
    seen: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        raw = row.get(key_col)
        if raw is None:
            continue
        key = str(raw).strip()
        if not key:
            continue
        seen.setdefault(key, []).append(row)
    index: dict[str, dict[str, Any]] = {}
    duplicates: set[str] = set()
    for key, matches in seen.items():
        if len(matches) == 1:
            index[key] = matches[0]
        else:
            duplicates.add(key)
    return index, duplicates


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv if argv is not None else sys.argv[1:])

    xlsx = Path(args.xlsx_path)
    if not xlsx.is_file():
        print(f"ERROR: not a file: {xlsx}")
        return 2

    if "template" not in args.list_name.lower() and not args.force:
        print(f"ERROR: list name '{args.list_name}' does not contain 'Template'. "
              f"Refusing to write to a non-template list without --force.")
        return 3

    field_allowlist = [f.strip() for f in args.fields.split(",") if f.strip()]
    if not field_allowlist:
        print("ERROR: --fields is empty; nothing to update.")
        return 4
    if args.key_col in field_allowlist:
        print(f"ERROR: --key-col '{args.key_col}' cannot also be in --fields "
              f"(would attempt to rewrite the natural key).")
        return 5

    try:
        from core.src.sharepoint_integration.config import GlobalSharePointConfig
        from core.src.sharepoint_integration.sp_session import SpSession
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: HILDA import failed: {type(exc).__name__}: {exc}")
        print("Run inside the hilda-worker container with PYTHONPATH=/app.")
        return 6

    cfg = GlobalSharePointConfig.from_sources(config_path=_resolve_config_path())
    if not cfg.username or not cfg.password:
        print("ERROR: NTLM credentials missing. Set them in "
              "config/sharepoint_integration.json or via HILDA_SP_* env vars.")
        return 7

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run this inside the hilda-worker container.")
        return 8

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        print("ERROR: empty worksheet -- no header row.")
        return 9

    header_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            header_map[idx] = name
    if not header_map:
        print("ERROR: header row had no non-empty cells.")
        return 10

    header_names = set(header_map.values())
    if args.key_col not in header_names:
        print(f"ERROR: key column '{args.key_col}' not found in Excel header row.")
        print(f"       Available columns: {sorted(header_names)}")
        return 11
    missing_fields = [f for f in field_allowlist if f not in header_names]
    if missing_fields:
        print(f"ERROR: --fields references columns not in Excel header row: {missing_fields}")
        print(f"       Available columns: {sorted(header_names)}")
        return 12

    print(f"[info] site_url   : {cfg.site_url}")
    print(f"[info] username   : {cfg.username}")
    print(f"[info] list_name  : {args.list_name}")
    print(f"[info] xlsx       : {xlsx}")
    print(f"[info] sheet      : {ws.title}")
    print(f"[info] key column : {args.key_col}")
    print(f"[info] fields     : {field_allowlist}")
    print(f"[info] filter     : {args.odata_filter or '(none -- reading ALL rows in list)'}")
    print(f"[info] mode       : {'DRY-RUN (no writes)' if args.dry_run else 'LIVE (SP writes)'}"
          + (" [always-send]" if args.always_send else " [diff-only]"))
    print()

    # Read SP first
    session = SpSession(
        site_url=cfg.site_url,
        ntlm_user=cfg.username,
        ntlm_pass=cfg.password,
    )
    print(f"[info] fetching SP rows from '{args.list_name}' ...")
    try:
        sp_rows, _warnings = _fetch_all_sp_rows(
            session, args.list_name, args.key_col, field_allowlist, args.odata_filter,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: SP read failed: {type(exc).__name__}: {exc}")
        return 13
    print(f"[info] SP rows fetched: {len(sp_rows)}")

    sp_index, dup_keys = _index_by_key(sp_rows, args.key_col)
    if dup_keys:
        print(f"[warn] {len(dup_keys)} key value(s) appear on multiple SP rows -- these "
              f"will be SKIPPED (ambiguous match). Narrow with --filter to fix.")
        print(f"       Ambiguous key values: {sorted(dup_keys)[:20]}"
              + ("..." if len(dup_keys) > 20 else ""))
    print(f"[info] SP rows indexed by unique key: {len(sp_index)}")
    print()

    # Iterate Excel data rows
    data_rows: list[tuple[int, tuple[Any, ...]]] = []
    for excel_row_num, row in enumerate(rows_iter, start=2):
        if all(cell is None for cell in row):
            continue
        data_rows.append((excel_row_num, row))
    if args.limit is not None:
        data_rows = data_rows[: args.limit]
    print(f"[info] Excel data rows to process: {len(data_rows)}")
    print()

    updated = 0
    no_diff = 0
    unmatched = 0
    ambiguous = 0
    skipped_no_key = 0
    failed = 0
    failed_rows: list[tuple[int, str]] = []

    # Reverse header lookup: col_name -> col_idx
    header_by_name = {name: idx for idx, name in header_map.items()}

    for i, (excel_row_num, row) in enumerate(data_rows, start=1):
        # 1. Extract key value
        key_idx = header_by_name[args.key_col]
        key_raw = row[key_idx] if key_idx < len(row) else None
        key_coerced = _coerce_cell(key_raw)
        if key_coerced is None or key_coerced == "":
            skipped_no_key += 1
            print(f"[skip] excel-row={excel_row_num}: empty {args.key_col}; skipping")
            continue
        key_str = str(key_coerced).strip()

        # 2. Ambiguous key?
        if key_str in dup_keys:
            ambiguous += 1
            print(f"[skip] excel-row={excel_row_num} {args.key_col}={key_str}: "
                  f"ambiguous on SP side (multiple matches); skipping")
            continue

        # 3. Look up SP row
        sp_row = sp_index.get(key_str)
        if sp_row is None:
            unmatched += 1
            print(f"[skip] excel-row={excel_row_num} {args.key_col}={key_str}: "
                  f"no matching SP row; skipping")
            continue

        # 4. Build MERGE payload with diff-only semantics
        payload: dict[str, Any] = {}
        for field in field_allowlist:
            col_idx = header_by_name[field]
            excel_val = _coerce_cell(row[col_idx] if col_idx < len(row) else None)
            sp_val = sp_row.get(field)
            if args.always_send:
                # send always -- but excel-empty -> skip (don't blank SP)
                if excel_val is not None:
                    payload[field] = excel_val
                continue
            # diff mode
            if excel_val is None:
                # Excel cell blank -> DON'T touch SP (preserves existing).
                continue
            if _norm_for_compare(excel_val) == _norm_for_compare(sp_val):
                continue  # no diff
            payload[field] = excel_val

        if not payload:
            no_diff += 1
            continue

        sp_id = sp_row["_sp_id"]
        if args.dry_run:
            print(f"[dry-run] excel-row={excel_row_num} {args.key_col}={key_str} "
                  f"sp_id={sp_id}: would MERGE "
                  f"{json.dumps(payload, default=str, sort_keys=True)}")
            updated += 1
        else:
            try:
                status = session.merge(
                    list_name=args.list_name,
                    customer_id="",
                    item_id=sp_id,
                    fields=payload,
                )
                if 200 <= status < 300:
                    updated += 1
                    if i % 10 == 0:
                        print(f"[progress] processed {i}/{len(data_rows)} "
                              f"(updated={updated} no_diff={no_diff} "
                              f"unmatched={unmatched} failed={failed})")
                else:
                    failed += 1
                    msg = f"HTTP {status}"
                    failed_rows.append((excel_row_num, msg))
                    print(f"[fail] excel-row={excel_row_num} sp_id={sp_id} {msg}")
            except Exception as exc:  # noqa: BLE001
                failed += 1
                msg = f"{type(exc).__name__}: {str(exc)[:180]}"
                failed_rows.append((excel_row_num, msg))
                print(f"[fail] excel-row={excel_row_num} sp_id={sp_id} {msg}")

    print()
    print(f"[done] updated={updated} no_diff={no_diff} "
          f"unmatched={unmatched} ambiguous={ambiguous} "
          f"skipped_no_key={skipped_no_key} failed={failed} "
          f"rows_processed={len(data_rows)}")
    if failed_rows:
        print("\nFAILED ROWS (excel-row-number, error):")
        for row_num, err in failed_rows:
            print(f"  row {row_num}: {err}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
