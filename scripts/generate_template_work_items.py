"""generate_template_work_items.py -- render an Excel row-set as YAML
work_items entries ready to paste into template.yaml.

Companion to create_sp_rows_from_excel.py -- same Excel file (row 1 =
internal column names, rows 2+ = per-row values). This script emits the
corresponding `template.yaml` work_items block so the two sources stay
consistent: whatever populated Deliverables_Template on SP now also seeds
HILDA's local template.yaml.

Reference structure (from customizations/template_schemas/MMK/template.yaml
work_items entry):

    - item_no: 2
      item_name: Sustainability Certification Form
      item_type: compliance_certification_release_notes
      tracking_modality: [Email]
      doc_count: 1
      review_required: false
      no_customer_upload: false
      force_tracking_enabled: true
      milestone_gating: true
      item_description: [["Sustainability"]]
      item_path_id: mno_cpm_item
      tg_path_id: mno_cpm
      target_folder: "Documentation/Certification/Sustainability Certification"
      handset: true
      tablet: true
      wearable: false
      ir: true
      osmr: true
      rmr: true
      hmr_smr: true
      tg_name: CPM
      ingress_nsd: None
      folder_routing_enabled: false
      tg_email_group_alias: null
      tg_owner_name: null
      tg_owner_corp_usa_email: null
      tg_owner_corp_email: null
      tg_owner_corp_id: null
      corp_id_list: null
      email_cc_list: null

USAGE:

    python scripts/generate_template_work_items.py \\
        --input rows.xlsx \\
        --output template_work_items.yaml

    # Optional flags:
    #   --sheet <name>      pick a non-default worksheet
    #   --wrap-key <name>   YAML section header (default: work_items)
    #   --no-wrap           emit bare list without the section header

TYPE COERCION RULES (per column):

  BOOL columns (review_required / no_customer_upload / force_tracking_enabled /
    milestone_gating / handset / tablet / wearable / ir / osmr / rmr /
    hmr_smr / folder_routing_enabled):
      TRUE / true / 1 / yes -> true;  FALSE / false / 0 / no -> false;
      empty cell -> false (SP default per FR-81)

  NUMERIC (item_no / doc_count / sort_order):
      integer as-is; empty cell -> null (SP default)

  LIST columns (tracking_modality / corp_id_list / email_cc_list):
      Cell value split on commas -> YAML inline list
      Empty cell -> null
      Cell literal `[X, Y]` -> passed through

  NESTED-LIST (item_description) per FR-82 tag semantics:
      Cell value `Sustainability` -> [["Sustainability"]]
      Cell with ';' separator (OR groups): `A; B, C` -> [["A"], ["B", "C"]]
      Cell literal `[[...]]` -> passed through
      Empty cell -> null

  PRESERVE-STRING (ingress_nsd / customer_delivery_modality):
      Cell value `None` -> literal string `None` (NOT YAML null; HILDA
      convention for these enum-like columns)
      Empty cell -> null

  QUOTED-STRING (target_folder / item_name):
      Wrap value in double quotes to survive YAML tokens like `/`, `:` etc.

  DEFAULT: string values passed through; empty -> null

If you need to override a column's category, edit the sets at the top of
the script.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Per-column categorization -- adjust here if headers differ or new columns
# need special handling.
# ---------------------------------------------------------------------------

_BOOL_COLS = frozenset({
    "review_required", "no_customer_upload", "force_tracking_enabled",
    "milestone_gating", "handset", "tablet", "wearable",
    "ir", "osmr", "rmr", "hmr_smr", "folder_routing_enabled",
})

_INT_COLS = frozenset({
    "item_no", "doc_count", "sort_order",
})

_LIST_COLS = frozenset({
    "tracking_modality", "corp_id_list", "email_cc_list",
})

_NESTED_LIST_COLS = frozenset({
    "item_description",
})

# Columns where the literal string "None" (or empty) means "no ingress /
# no modality" -- HILDA convention, NOT YAML null (used for enum-like
# fields where the enum has a None member).
_PRESERVE_STRING_COLS = frozenset({
    "ingress_nsd", "customer_delivery_modality",
})

# Columns whose values benefit from being quoted (contain '/', ':' etc.)
_QUOTED_STRING_COLS = frozenset({
    "target_folder", "item_name", "customer_delivery_info",
})


_TRUTHY = frozenset({"true", "True", "TRUE", "yes", "YES", "1", 1, True})
_FALSY  = frozenset({"false", "False", "FALSE", "no", "NO", "0", 0, False})


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _coerce_bool(v: Any, default: bool = False) -> bool:
    if _is_empty(v):
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return bool(v)
    s = str(v).strip()
    if s in _TRUTHY:
        return True
    if s in _FALSY:
        return False
    # Default to False for unknown -- log via a bare comment upstream would be nice
    return default


def _coerce_int(v: Any) -> int | None:
    if _is_empty(v):
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        try:
            return int(float(v))  # tolerates "3.0"
        except Exception:  # noqa: BLE001
            return None


def _yaml_quote(s: str) -> str:
    """Return YAML double-quoted form; escape backslash + double-quote."""
    escaped = s.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def _needs_quoting(s: str) -> bool:
    """Return True if the bare string would confuse a YAML parser."""
    if s == "":
        return True
    if s.strip() != s:
        return True
    # YAML reserved / structural chars
    if any(c in s for c in [":", "#", "&", "*", "!", "|", ">", "'", '"', "%", "@", "`"]):
        return True
    # Starts with a token that YAML would parse as something other than a bare string
    if s.strip().lower() in {"null", "~", "true", "false", "yes", "no", "on", "off"}:
        return True
    if re.match(r"^[-+]?\d", s):
        # Looks numeric; quote to force string
        return True
    if s.startswith("[") or s.startswith("{"):
        # Looks like flow YAML
        return True
    return False


def _emit_bare_or_quoted(v: str) -> str:
    """Emit a scalar string as bare-YAML if safe, else quoted."""
    if _needs_quoting(v):
        return _yaml_quote(v)
    return v


def _emit_list_inline(items: list[str]) -> str:
    """Emit as flow-style YAML list [a, b, c] with per-item bare/quoted."""
    parts = [_emit_bare_or_quoted(str(x).strip()) for x in items if str(x).strip()]
    return "[" + ", ".join(parts) + "]"


def _emit_nested_list_inline(groups: list[list[str]]) -> str:
    """Emit as flow-style YAML list-of-lists [["a"], ["b", "c"]].
    Per HILDA convention (reference: template.yaml item_description entries),
    inner strings are ALWAYS double-quoted for uniformity."""
    outer_parts = []
    for group in groups:
        inner = ", ".join(_yaml_quote(str(x).strip()) for x in group if str(x).strip())
        outer_parts.append("[" + inner + "]")
    return "[" + ", ".join(outer_parts) + "]"


def _coerce_list_cell(cell: Any) -> list[str] | None:
    """CSV cell -> list; already-list literal `[a, b]` passed through as parsed;
    empty -> None."""
    if _is_empty(cell):
        return None
    s = str(cell).strip()
    # If already looks like flow YAML list, try to parse minimally
    if s.startswith("[") and s.endswith("]"):
        inner = s[1:-1].strip()
        if not inner:
            return []
        return [p.strip() for p in inner.split(",") if p.strip()]
    # Otherwise treat as comma-separated
    return [p.strip() for p in s.split(",") if p.strip()]


def _coerce_nested_list_cell(cell: Any) -> list[list[str]] | None:
    """FR-82 item_description parsing.
      Cell `Sustainability`               -> [["Sustainability"]]
      Cell `A, B`                          -> [["A", "B"]]     (AND group)
      Cell `A; B, C`                       -> [["A"], ["B", "C"]] (OR of AND-groups)
      Cell literal `[["A"]]`                -> passed through
    Empty -> None."""
    if _is_empty(cell):
        return None
    s = str(cell).strip()
    # Already looks like list-of-lists literal? crude parse.
    if s.startswith("[[") and s.endswith("]]"):
        # Strip outer brackets: `[[a], [b, c]]` -> `[a], [b, c]`
        inner_body = s[2:-2].strip()
        # Split on `],` boundaries
        segments = re.split(r"\]\s*,\s*\[", inner_body)
        groups = []
        for seg in segments:
            seg = seg.strip("[]").strip()
            if not seg:
                continue
            groups.append([p.strip().strip('"').strip("'") for p in seg.split(",") if p.strip()])
        return groups or None
    # Split on `;` for OR groups; each group on `,` for AND
    if ";" in s:
        return [
            [p.strip() for p in group.split(",") if p.strip()]
            for group in s.split(";")
            if group.strip()
        ]
    # Single AND group
    return [[p.strip() for p in s.split(",") if p.strip()]]


def _emit_scalar_value(col_name: str, cell: Any) -> str:
    """Return the right-hand-side YAML string for one column's value."""
    # BOOL columns
    if col_name in _BOOL_COLS:
        return "true" if _coerce_bool(cell) else "false"

    # INT columns
    if col_name in _INT_COLS:
        v = _coerce_int(cell)
        return "null" if v is None else str(v)

    # LIST columns -- inline flow
    if col_name in _LIST_COLS:
        parsed = _coerce_list_cell(cell)
        if parsed is None:
            return "null"
        if not parsed:
            return "[]"
        return _emit_list_inline(parsed)

    # NESTED-LIST columns
    if col_name in _NESTED_LIST_COLS:
        parsed = _coerce_nested_list_cell(cell)
        if parsed is None:
            return "null"
        if not parsed:
            return "[]"
        return _emit_nested_list_inline(parsed)

    # PRESERVE-STRING columns -- "None" stays literal
    if col_name in _PRESERVE_STRING_COLS:
        if _is_empty(cell):
            return "null"
        s = str(cell).strip()
        # HILDA convention: bare `None` preserved (not YAML null)
        if s == "None":
            return "None"
        return _emit_bare_or_quoted(s)

    # QUOTED-STRING columns
    if col_name in _QUOTED_STRING_COLS:
        if _is_empty(cell):
            return "null"
        return _yaml_quote(str(cell).strip())

    # DEFAULT: string or null
    if _is_empty(cell):
        return "null"
    s = str(cell).strip()
    return _emit_bare_or_quoted(s)


def _parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n", 1)[0])
    p.add_argument("--input", required=True, help="Path to .xlsx (row 1 = column names).")
    p.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet).")
    p.add_argument("--output", default="-", help="Output YAML path. '-' = stdout.")
    p.add_argument("--wrap-key", default="work_items",
                   help="Top-level YAML key to wrap the list under (default: work_items). "
                        "Use with --no-wrap to disable wrapping entirely.")
    p.add_argument("--no-wrap", action="store_true",
                   help="Emit the bare list of entries without a wrapping section header.")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv if argv is not None else sys.argv[1:])

    xlsx = Path(args.input)
    if not xlsx.is_file():
        print(f"ERROR: not a file: {xlsx}", file=sys.stderr)
        return 2

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run inside hilda-worker container "
              "OR pip install openpyxl on host.", file=sys.stderr)
        return 3

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        print("ERROR: worksheet is empty.", file=sys.stderr)
        return 4

    # Header index -> column name (skip blank header cells)
    headers: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            headers.append((idx, name))
    if not headers:
        print("ERROR: header row had no non-empty column names.", file=sys.stderr)
        return 5

    out_lines: list[str] = []
    out_lines.append(f"# Generated by scripts/generate_template_work_items.py")
    out_lines.append(f"# Source: {xlsx}")
    out_lines.append(f"# Sheet : {ws.title}")
    out_lines.append(f"# Columns detected ({len(headers)}): "
                     + ", ".join(name for _, name in headers))
    out_lines.append(f"#")
    out_lines.append(f"# Paste under `work_items:` in "
                     f"customizations/template_schemas/<customer_id>/template.yaml")
    out_lines.append("")

    if not args.no_wrap:
        out_lines.append(f"{args.wrap_key}:")
        out_lines.append("")

    row_indent = "  " if not args.no_wrap else ""
    field_indent = row_indent + "  "

    row_count = 0
    for excel_row_num, row in enumerate(rows_iter, start=2):
        if all(cell is None for cell in row):
            continue
        row_count += 1
        first_field = True
        for col_idx, col_name in headers:
            cell_value = row[col_idx] if col_idx < len(row) else None
            rhs = _emit_scalar_value(col_name, cell_value)
            prefix = f"{row_indent}- " if first_field else f"{field_indent}"
            out_lines.append(f"{prefix}{col_name}: {rhs}")
            first_field = False
        out_lines.append("")  # blank line between items

    body = "\n".join(out_lines).rstrip() + "\n"

    if args.output == "-":
        sys.stdout.write(body)
    else:
        Path(args.output).write_text(body, encoding="utf-8")
        print(f"[done] wrote {row_count} work_items to {args.output}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
